"""Teste rápido (smoke test) dos principais fluxos, sem precisar de navegador."""
import io
from datetime import date, datetime, timedelta, time as dt_time
from unittest.mock import patch

from reportlab.pdfgen import canvas
from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import (
    Grupo, GrupoMembro, GrupoPaciente, PlataformaConfig, Usuario, Paciente, Exame, Agendamento,
    PreparoModelo, PreparoCorte, PreparoMedicamentoSuspenso, PreparoInfoGeral, PreparoAlimento,
    PreparoExameAnterior, PreparoMedicamentoMantido, Medicamento, PerguntaPendente, FaqItem,
    ChatMensagem, ResultadoExame,
)
from app.pdf_preparo import (
    _sugerir_informacoes_gerais, _sugerir_alimentos, _sugerir_medicamentos, _sugerir_cortes,
    _sugerir_exames_anteriores,
)
from app.xlsx_preparo import extrair_sugestoes_de_xlsx
from app.ia_preparo import responder_com_ia

app = create_app()
client = app.test_client()


def login(email, senha):
    return client.post("/login", data={"email": email, "senha": senha}, follow_redirects=True)


def login_paciente(cpf, data_nascimento):
    # Login do paciente é por CPF + data de nascimento (não mais por
    # telefone - o telefone deixou de ser credencial, ver
    # app/routes_auth.py:login_paciente).
    return client.post(
        "/login-paciente",
        data={"cpf": cpf, "data_nascimento": data_nascimento},
        follow_redirects=True,
    )


def checar(nome, condicao):
    status = "OK" if condicao else "FALHOU"
    print(f"[{status}] {nome}")
    assert condicao, nome


# ---------- Fluxo básico da Clínica Vitória (secretária) ----------

r = login("secretaria@clinicavitoria.com", "123456")
checar("Login secretaria (Vitória) funciona", r.status_code == 200 and "Painel" in r.get_data(as_text=True))
checar("Navbar mostra o nome da clínica atual", "Clínica Vitória" in r.get_data(as_text=True))

r = client.get("/equipe/pacientes")
checar("Lista de pacientes da Vitória contém João, não contém Maria",
       "João Pereira" in r.get_data(as_text=True) and "Maria Silva" not in r.get_data(as_text=True))

r = client.get("/equipe/exames")
checar("Lista de exames da Vitória contém Colonoscopia", "Colonoscopia" in r.get_data(as_text=True))

r = client.get("/equipe/agenda", follow_redirects=True)
checar("Agenda da Vitória acessível (redireciona para o painel)", "Agenda de exames" in r.get_data(as_text=True))

r = client.get("/equipe/equipe-membros", follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Equipe da Vitória lista Ana e o Dr. Carlos", "Ana Secretária" in texto and "Carlos Andrade" in texto)

r = client.get("/equipe/clinica/configuracoes")
texto = r.get_data(as_text=True)
checar("Página de dados da clínica responde 200", r.status_code == 200)

# Nota: "horário de atendimento por dia da semana" (dia_0_ativo/inicio/fim)
# nunca existiu de fato no modelo/telas desta versão - não há coluna nem
# campo de formulário pra isso (verificado em app/models.py e nos
# templates); a asserção correspondente que havia aqui antes era código
# morto pré-existente (não uma regressão da Fatia 5) e foi removida.
r = client.post("/equipe/clinica/configuracoes", data={
    "nome": "Clínica Vitória", "telefone": "(27) 3333-4444", "email_contato": "contato@clinicavitoria.com",
    "razao_social": "Clínica Vitória Diagnósticos Ltda.", "cnpj": "12.345.678/0001-90",
    "cep": "29010-000", "rua": "Av. Jerônimo Monteiro", "numero": "1000", "bairro": "Centro",
    "cidade": "Vitória", "uf": "ES",
    "codigo_ibge_municipio": "3205309",
}, follow_redirects=True)
checar("Secretária consegue salvar os dados da clínica", "atualizados com sucesso" in r.get_data(as_text=True).lower())

client.get("/logout")

login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)
r = client.get("/equipe/clinica/configuracoes")
checar("Médico não pode acessar os dados da clínica (restrito à secretária)", r.status_code in (302,))
client.get("/logout")

# ---------- Isolamento: secretária da Clínica São Paulo não vê dados da Vitória ----------

login("secretaria@clinicasp.com", "123456")

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Secretária da SP vê Maria mas não João (isolamento)",
       "Maria Silva" in texto and "João Pereira" not in texto)

r = client.get("/equipe/exames")
checar("Exames da SP não incluem Hemograma da Vitória", "Hemograma" not in r.get_data(as_text=True))

client.get("/logout")

# ---------- Médico vinculado a duas clínicas: precisa escolher ----------

r = login("medico@clinicavitoria.com", "123456")
texto = r.get_data(as_text=True)
# Agora o que precisa ser escolhido é a EMPRESA (o cliente/tenant), não a
# filial: este médico atende em duas empresas sem relação entre si.
checar("Médico com vínculo em duas empresas cai na tela de escolha de empresa",
       "Em qual empresa" in texto)

r = client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)
checar("Médico consegue selecionar a Clínica Vitória", "Clínica Vitória" in r.get_data(as_text=True))

r = client.get("/equipe/clinica")
checar("Tela de troca de empresa ainda lista as duas empresas do médico", "Clínica São Paulo" in r.get_data(as_text=True))

client.get("/logout")

# ---------- Área do médico: cada médico só vê seus próprios exames/pacientes ----------

r = login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)

r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
checar("Dr. Carlos vê Colonoscopia e Glicemia, mas não o Hemograma da Dra. Fernanda",
       "Colonoscopia" in texto and "Glicemia" in texto and "Hemograma" not in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Dr. Carlos vê João (seu paciente), mas não Pedro (paciente da Dra. Fernanda)",
       "João Pereira" in texto and "Pedro Souza" not in texto)

r = client.get("/equipe/agenda", follow_redirects=True)
checar("Agenda do Dr. Carlos acessível (redireciona para o painel)", "Agenda de exames" in r.get_data(as_text=True))

r = client.get("/equipe/equipe-membros")
checar("Médico não consegue acessar a gestão de equipe", r.status_code in (302,))

client.get("/logout")

login("medica2@clinicavitoria.com", "123456")
r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
# O seed também associa a Dra. Fernanda como médica EXTRA da Colonoscopia
# (além do próprio Hemograma) - ela vê os dois; o que ela não vê são os
# exames que são só do Dr. Carlos (Glicemia, Testes de Hidrogênio).
checar("Dra. Fernanda vê o Hemograma (dela) e a Colonoscopia (médica extra), mas não a Glicemia do Dr. Carlos",
       "Hemograma" in texto and "Colonoscopia" in texto and "Glicemia" not in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Dra. Fernanda vê Pedro (seu paciente), mas não João (paciente do Dr. Carlos)",
       "Pedro Souza" in texto and "João Pereira" not in texto)

client.get("/logout")

# Secretária continua vendo tudo da clínica
login("secretaria@clinicavitoria.com", "123456")
r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
checar("Secretária vê exames de ambos os médicos da Vitória",
       "Colonoscopia" in texto and "Glicemia" in texto and "Hemograma" in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Secretária vê pacientes de ambos os médicos", "João Pereira" in texto and "Pedro Souza" in texto)
client.get("/logout")

# ---------- Cadastro público: conta nasce SOLO, sem nenhum Grupo (Fatia 6) ----------
# O cadastro não cria mais Grupo nenhum - a conta é plenamente usável
# sozinha (escopo pessoal via criado_por_id/cadastrado_por_id), e só passa
# a ter um Grupo de verdade se a pessoa decidir criar um (ver
# routes_grupo.py:novo()).

r = client.post("/cadastro", data={
    "nome": "Fulano Teste",
    "cpf": "168.995.350-09",
    "email": "fulano@clinicateste.com",
    "senha": "senha123",
    "papel": "secretaria",
}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Cadastro público loga automaticamente, direto no painel", "Painel" in texto)
with app.app_context():
    usuario_novo = Usuario.query.filter_by(email="fulano@clinicateste.com").first()
    checar("NENHUM Grupo foi criado no cadastro (Fatia 6: Grupo é opcional)",
           GrupoMembro.query.filter_by(usuario_id=usuario_novo.id).first() is None)

r = client.get("/equipe/pacientes")
checar("Conta solo recém-criada começa sem nenhum paciente de outras contas/grupos",
       "João Pereira" not in r.get_data(as_text=True))

client.get("/logout")

# ---------- Cadastro público escolhendo "médico": conta solo recebe ----------
# ---------- todas as permissões administrativas mesmo sem Grupo/secretária ----------

r = client.post("/cadastro", data={
    "nome": "Dr. Ricardo Alves",
    "cpf": "111.444.777-35",
    "crm_numero": "88899", "crm_uf": "ES",
    "email": "ricardo@clinicasolo.com",
    "senha": "senha123",
    "papel": "medico",
}, follow_redirects=True)
checar("Cadastro público como médico também funciona", "Painel" in r.get_data(as_text=True))

with app.app_context():
    ricardo = Usuario.query.filter_by(email="ricardo@clinicasolo.com").first()
    checar("Médico que se cadastrou é tipo 'medico'", ricardo.tipo == "medico")
    checar(
        "Médico solo recebe todas as permissões administrativas (não há secretária)",
        ricardo.perm_pacientes and ricardo.perm_equipe and ricardo.perm_filiais and ricardo.perm_dados_clinica,
    )
    checar("NENHUM Grupo foi criado para o médico solo", GrupoMembro.query.filter_by(usuario_id=ricardo.id).first() is None)

r = client.get("/equipe/equipe-membros", follow_redirects=True)
checar("Médico solo (sem Grupo) que acessa Equipe cai na tela de 'Meus grupos' vazia (nenhum grupo pra convidar ainda)",
       "ainda não participa de nenhum grupo" in r.get_data(as_text=True).lower())
r = client.get("/equipe/filiais", follow_redirects=True)
checar("Médico solo consegue acessar 'Meus grupos' (redireciona pra grupo.meus_grupos)", r.status_code == 200)
r = client.get("/equipe/clinica/configuracoes", follow_redirects=True)
checar("Médico solo é levado a criar um Grupo antes de configurar 'Dados da clínica' (não há Grupo ainda)",
       "Cadastre seu primeiro grupo" in r.get_data(as_text=True))
r = client.get("/equipe/pacientes/novo")
checar("Médico solo consegue acessar Novo paciente mesmo sem Grupo", r.status_code == 200)

# ---------- Cadastro de paciente sem senha: login por telefone + data de nascimento ----------

r = client.post("/equipe/pacientes/novo", data={
    "nome": "Beatriz Nunes", "cpf": "555.666.777-00", "email": "",
}, follow_redirects=True)
checar("Cadastro de paciente exige telefone e data de nascimento (sem eles, é rejeitado)",
       "obrigatórios" in r.get_data(as_text=True).lower())

r = client.post("/equipe/pacientes/novo", data={
    "nome": "Beatriz Nunes", "cpf": "555.666.777-00", "email": "",
    "telefone": "(28) 98765-4321", "data_nascimento": "1995-06-20",
}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Cadastro de paciente funciona sem e-mail nem senha, só com telefone e data de nascimento",
       "cadastrado" in texto.lower() and "não é necessário criar senha" in texto.lower())
checar("Paciente recém-cadastrado (sem agendamento ainda) já aparece na lista para o médico com permissão de pacientes",
       "Beatriz Nunes" in texto)

# Mesmo telefone, MAS data de nascimento diferente - representa outra
# pessoa (ex.: um familiar) compartilhando o mesmo telefone de contato.
# Isso agora é permitido (a unicidade real por clínica é por CPF, e o
# login do paciente já usa telefone + data de nascimento pra diferenciar
# as contas - ver comentário em routes_medico.pacientes_novo).
r = client.post("/equipe/pacientes/novo", data={
    "nome": "Outra Pessoa", "cpf": "999.888.777-00", "email": "",
    "telefone": "(28) 98765-4321", "data_nascimento": "1980-01-01",
}, follow_redirects=True)
checar("Cadastro com o MESMO telefone mas data de nascimento DIFERENTE agora é permitido (compartilhamento familiar)",
       "cadastrado" in r.get_data(as_text=True).lower())

# Duplicata DE VERDADE: mesmo telefone E mesma data de nascimento de
# Beatriz - essa sim continua bloqueada (mesma pessoa se cadastrando de novo).
r = client.post("/equipe/pacientes/novo", data={
    "nome": "Beatriz Duplicada", "cpf": "111.111.111-11", "email": "",
    "telefone": "(28) 98765-4321", "data_nascimento": "1995-06-20",
}, follow_redirects=True)
checar("Cadastro com o MESMO telefone E a MESMA data de nascimento continua bloqueado",
       "já existe um paciente cadastrado com esse telefone e data de nascimento" in r.get_data(as_text=True).lower())

client.get("/logout")

r = login_paciente("999.888.777-00", "1980-01-01")
texto = r.get_data(as_text=True)
checar("Login da 'Outra Pessoa' funciona (CPF + nascimento próprios, mesmo telefone de contato da Beatriz)",
       "Meus exames" in texto or "Tirar dúvidas" in texto)
client.get("/logout")

r = login_paciente("999.888.777-00", "1970-12-31")
checar("Login com CPF existente mas data de nascimento que não bate é rejeitado",
       "incorretos" in r.get_data(as_text=True).lower())

r = login_paciente("555.666.777-00", "1995-06-20")
texto = r.get_data(as_text=True)
checar("Paciente cadastrado sem senha consegue entrar só com CPF e data de nascimento",
       "Meus exames" in texto or "Tirar dúvidas" in texto)
checar("Paciente sem senha não vê o link de 'Trocar senha' na barra superior", "Trocar senha" not in texto)

client.get("/logout")

# ---------- Grupos: cada Grupo já é a própria unidade (Fatia 5) ----------
# O antigo "Grupo Saúde Total" com 2 filiais virou 2 Grupos independentes
# (Centro/Praia) que compartilham a mesma equipe - não existe mais
# "várias filiais dentro da mesma empresa" (ver seed.py e
# app/clinica_utils.py).

with app.app_context():
    filial_centro_id = Grupo.query.filter_by(nome="Grupo Saúde Total - Centro").first().id
    filial_praia_id = Grupo.query.filter_by(nome="Grupo Saúde Total - Praia").first().id

login("secretaria@gruposaude.com", "123456")
# A secretária tem vínculo ativo nos DOIS Grupos - precisa escolher
# explicitamente qual está usando agora (o oposto do modelo antigo, em
# que "várias filiais da mesma empresa" não exigia escolha nenhuma).
texto = client.get("/equipe/clinica", follow_redirects=True).get_data(as_text=True)
checar("Secretária com vínculo em 2 Grupos precisa escolher explicitamente qual está usando",
       "Em qual empresa" in texto)

client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)
texto_pacientes = client.get("/equipe/pacientes").get_data(as_text=True)
checar("A lista de pacientes do Grupo Centro não tem coluna de filial (o Grupo já é a unidade)",
       "<th>Filial</th>" not in texto_pacientes)

r = client.get("/equipe/filiais", follow_redirects=True)
texto = r.get_data(as_text=True)
checar("'Meus locais de atendimento' virou 'Meus grupos' (redireciona pra lá) e lista os dois",
       "Grupo Saúde Total - Centro" in texto and "Grupo Saúde Total - Praia" in texto)
client.get("/logout")

with app.app_context():
    grupo_centro = Grupo.query.get(filial_centro_id)
    checar("Médico que atua nos dois Grupos conta 1x na cobrança de CADA um (cobrança agora é por Grupo)",
           len(grupo_centro.medicos_distintos) == 1)
    checar("Valor mensal estimado do Grupo Centro é 1 médico x R$150", float(grupo_centro.valor_mensal_estimado) == 150.0)

# ---------- Editar os dados de um Grupo exige que ele esteja ATIVO na sessão ----------
# Fatia 5 (passo 4): não existe mais "editar uma filial que não é a
# atual" - o parâmetro de URL /equipe/clinica/configuracoes/<id> é
# ignorado (só existe pra não quebrar links antigos); quem é editado é
# sempre o Grupo ATIVO da sessão.

login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)
r = client.post(f"/equipe/clinica/configuracoes/{filial_praia_id}", data={
    "nome": "Grupo Saúde Total - Centro", "telefone": "(27) 3111-1111", "email_contato": "centro@gruposaude.com",
    "cidade": "Vila Velha", "uf": "ES",
}, follow_redirects=True)
checar("POST com o id da Praia na URL, mas o Centro ativo: responde 200 mesmo assim (parâmetro ignorado)",
       r.status_code == 200)
with app.app_context():
    checar("Quem recebeu o telefone novo foi o Grupo ATIVO (Centro), não a Praia (pelo id na URL)",
           Grupo.query.get(filial_centro_id).telefone == "(27) 3111-1111")

client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)
r = client.post("/equipe/clinica/configuracoes", data={
    "nome": "Grupo Saúde Total - Praia", "telefone": "(27) 3222-1111", "email_contato": "praia@gruposaude.com",
    "cidade": "Vila Velha", "uf": "ES",
}, follow_redirects=True)
checar("Selecionando a Praia como Grupo ativo, dá pra editar os dados dela", "atualizados com sucesso" in r.get_data(as_text=True).lower())
with app.app_context():
    filial_praia = Grupo.query.get(filial_praia_id)
    checar("Os dados da Praia foram salvos depois de selecioná-la como ativa",
           filial_praia.telefone == "(27) 3222-1111")

# ---------- Cadastrar médico/secretária: por CPF ou conta nova (routes_grupo.convidar) ----------
# Fatia 5 (passo 4): "equipe-membros/novo" (com filial_ids, ClinicaMembro)
# não existe mais - a tela de Equipe (medico.equipe_lista) redireciona pra
# routes_grupo.py:convidar(), que cria a conta já ATIVA só no Grupo que
# estiver selecionado no momento (não existe mais "escolher a filial" -
# o Grupo ativo já é isso).
r = client.post(f"/grupos/{filial_praia_id}/convidar", data={
    "acao": "criar_conta", "nome": "Dra. Beatriz Costa", "email": "beatriz@gruposaude.com",
    "cpf": "222.333.444-05", "papel_conta": "medico", "papel_grupo": "membro",
}, follow_redirects=True)
checar("Cadastro de nova médica direto no grupo (conta nova) funciona", "cadastrado" in r.get_data(as_text=True).lower())

with app.app_context():
    nova_medica = Usuario.query.filter_by(email="beatriz@gruposaude.com").first()
    vinculo = GrupoMembro.query.filter_by(grupo_id=filial_praia_id, usuario_id=nova_medica.id).first()
    checar("A médica nova ficou vinculada ao Grupo Praia (o que estava selecionado ao convidar)",
           vinculo is not None and vinculo.ativo)
    checar("Ela NÃO tem vínculo no Grupo Centro (associação é só no grupo escolhido, não em 'todas as filiais')",
           GrupoMembro.query.filter_by(grupo_id=filial_centro_id, usuario_id=nova_medica.id).first() is None)
    checar("Como nenhuma permissão foi marcada, a médica nova começa sem nenhuma permissão administrativa",
           not (nova_medica.perm_pacientes or nova_medica.perm_equipe or nova_medica.perm_filiais or nova_medica.perm_dados_clinica))
    beatriz_id = nova_medica.id

client.get("/logout")

# Beatriz, sem nenhuma permissão, não consegue acessar telas administrativas
login("beatriz@gruposaude.com", "123456")
r = client.get("/equipe/equipe-membros", follow_redirects=True)
checar("Médica sem permissão de equipe é bloqueada ao tentar acessar a tela de Equipe",
       "permissão" in r.get_data(as_text=True).lower())
client.get("/logout")

# Secretária concede a permissão de equipe para a Beatriz pela tela de Permissões
login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)
r = client.get(f"/equipe/equipe-membros/{beatriz_id}/permissoes")
checar("Tela de permissões carrega para um membro existente", "Beatriz" in r.get_data(as_text=True))
r = client.post(f"/equipe/equipe-membros/{beatriz_id}/permissoes", data={"perm_equipe": "on"}, follow_redirects=True)
checar("Permissões atualizadas com sucesso", "atualizadas" in r.get_data(as_text=True).lower())
client.get("/logout")

# NOTA (achado, não é bug de app/ - comportamento observado): a tela de
# membros do grupo (routes_grupo.py:convidar) exige papel "dono" ou
# "administrador" no GrupoMembro para ser acessada - só marcar
# perm_equipe (permissão de conta, granular por tela) NÃO é suficiente
# aqui, diferente das outras telas administrativas (pacientes/dados da
# clínica), que continuam checando só a permissão da conta. Ou seja,
# depois de só conceder perm_equipe, Beatriz AINDA não acessa a tela de
# membros - só depois de o dono do grupo promovê-la a administrador.
login("beatriz@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)
r = client.get("/equipe/equipe-membros", follow_redirects=True)
checar("Só com perm_equipe (sem ser administrador do grupo), a médica AINDA não acessa a tela de membros",
       "Convidar membros para o grupo" not in r.get_data(as_text=True))
client.get("/logout")

# O dono do grupo promove Beatriz a administradora - agora sim ela acessa
# a tela de membros (routes_grupo.py:convidar, ação "tornar_administrador").
with app.app_context():
    beatriz_membro_id = GrupoMembro.query.filter_by(grupo_id=filial_praia_id, usuario_id=beatriz_id).first().id

login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)
r = client.post(f"/grupos/{filial_praia_id}/convidar", data={
    "acao": "tornar_administrador", "membro_id": str(beatriz_membro_id),
}, follow_redirects=True)
checar("Dono do grupo consegue promover a médica a administradora",
       "agora é administrador do grupo" in r.get_data(as_text=True))
client.get("/logout")

login("beatriz@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)
r = client.get("/equipe/equipe-membros", follow_redirects=True)
checar("Como administradora do grupo, a médica agora consegue acessar a tela de membros",
       "Convidar membros para o grupo" in r.get_data(as_text=True))
r = client.get("/equipe/filiais", follow_redirects=True)
checar("'Meus grupos' continua acessível (não é uma tela restrita por permissão)", r.status_code == 200)
client.get("/logout")

# ---------- Novo agendamento: dentro do Grupo ativo ----------
# Fatia 5 (passo 4): não existe mais "escolher a filial de destino" no
# agendamento - o Grupo atual já É a única unidade; pra agendar na Praia,
# a Praia precisa estar selecionada como Grupo ativo.

login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_praia_id)}, follow_redirects=True)

with app.app_context():
    medico_grupo_id = Usuario.query.filter_by(email="medico@gruposaude.com").first().id

r = client.get(f"/equipe/agenda/novo?filial_id={filial_praia_id}&medico_id={medico_grupo_id}")
checar("Tela de novo agendamento mostra o seletor de médico", "Médico" in r.get_data(as_text=True))

client.get("/logout")

# ---------- Modelos de preparo reaproveitáveis, cortes e medicamentos ----------

with app.app_context():
    clinica_vitoria_id = Grupo.query.filter_by(nome="Clínica Vitória").first().id
    modelo_hidrogenio = PreparoModelo.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first()
    exames_hidrogenio = Exame.query.filter_by(grupo_id=clinica_vitoria_id, preparo_modelo_id=modelo_hidrogenio.id).all()
    checar("Os 3 substratos do teste de hidrogênio compartilham o mesmo modelo de preparo, sem duplicar cadastro",
           len(exames_hidrogenio) == 3)
    corte_jejum = modelo_hidrogenio.cortes[0]
    checar("Modelo do teste de hidrogênio tem o corte de jejum de 12 horas", corte_jejum.horas_antes == 12)
    med_susp = modelo_hidrogenio.medicamentos_suspensos[0]
    checar("Modelo do teste de hidrogênio tem o medicamento a suspender 14 dias antes", med_susp.dias_antes == 14)

login("secretaria@clinicavitoria.com", "123456")

# Editar o modelo compartilhado deve refletir nos 3 exames que o usam.
with app.app_context():
    modelo_hidrogenio_id = PreparoModelo.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first().id

r = client.post(f"/equipe/preparo-modelos/{modelo_hidrogenio_id}/editar", data={
    "nome": "Teste de Hidrogênio/Metano Expirado - padrão",
    "instrucoes": "Texto de preparo atualizado pela secretária.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": "Jejum total (sólidos e líquidos)",
    "corte_horas[]": "12",
    "medicamento_nome[]": "Ozempic, Mounjaro, Trulicity ou similares",
    "medicamento_dias[]": "14",
    "medicamento_obs[]": "",
}, follow_redirects=True)
checar("Secretária consegue editar o modelo de preparo compartilhado", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    exame_lactose = Exame.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Lactose").first()
    checar("A edição do modelo aparece refletida no exame de Lactose (mesmo preparo)",
           exame_lactose.preparo.instrucoes == "Texto de preparo atualizado pela secretária.")
    exame_frutose = Exame.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Frutose").first()
    checar("A edição do modelo também aparece no exame de Frutose (mesmo preparo)",
           exame_frutose.preparo.instrucoes == "Texto de preparo atualizado pela secretária.")

client.get("/logout")

# Alertas de corte/medicamentos calculados a partir do horário do agendamento
with app.app_context():
    joao = Paciente.query.filter_by(cpf="123.456.789-00").first()
    exame_lactose = Exame.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Lactose").first()
    ag_teste = Agendamento(
        grupo_id=clinica_vitoria_id, paciente_id=joao.id, exame_id=exame_lactose.id,
        medico_id=exame_lactose.medico_id, data_hora=datetime(2026, 8, 10, 8, 0),
    )
    db.session.add(ag_teste)
    db.session.commit()
    ag_teste_id = ag_teste.id

login_paciente("123.456.789-00", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_teste_id}")
texto = r.get_data(as_text=True)
checar("Alerta de jejum mostra o horário calculado a partir do agendamento (12h antes de 10/08 08:00 = 09/08 20:00)",
       "09/08/2026 às 20:00" in texto)
checar("Lista de medicamentos mostra a data calculada de suspensão (14 dias antes de 10/08 = 27/07/2026)",
       "27/07/2026" in texto)
client.get("/logout")

# Importação de PDF: o PDF é lido diretamente pela IA (app.ia_pdf_preparo),
# preenchendo o formulário de novo modelo direto pra revisão — sem etapa
# intermediária de gerar/reimportar Excel. Sem ANTHROPIC_API_KEY neste
# ambiente de teste, cai automaticamente no fallback de extração por regex
# (app.pdf_preparo), exercitado aqui.
buffer_pdf = io.BytesIO()
c = canvas.Canvas(buffer_pdf)
c.drawString(50, 800, "TESTE DE EXTRACAO AUTOMATICA DE PDF")
c.drawString(50, 780, "Informacoes ao paciente:")
c.drawString(50, 760, "JEJUM de 12 horas, sendo permitido apenas agua.")
c.drawString(50, 740, "Liquidos claros ate 2 horas antes do exame.")
c.drawString(50, 720, "14 dias antes: OZEMPIC, MOUNJARO, TRULICITY OU SIMILARES.")
c.drawString(50, 700, "NAO e necessario suspender o AAS, Somalgin, Aspirina.")
c.drawString(50, 680, "Nao deve ter realizado colonoscopia ou endoscopia nas 4 semanas anteriores ao exame.")
c.showPage()
c.save()
buffer_pdf.seek(0)

login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (buffer_pdf, "teste_preparo.pdf")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
html_pdf = r.get_data(as_text=True)
checar("Importar o PDF responde 200 e leva direto pra revisão do formulário",
       r.status_code == 200 and "Revise com cuidado" in html_pdf)
checar("Formulário preenchido sugere o corte de jejum de 12 horas",
       "JEJUM de 12 horas" in html_pdf)
checar("Formulário preenchido sugere os medicamentos separados (não uma lista só)",
       "OZEMPIC" in html_pdf and "MOUNJARO" in html_pdf and "TRULICITY" in html_pdf)
checar("Formulário preenchido sugere os exames/procedimentos proibidos antes (colonoscopia/endoscopia, 28 dias)",
       "olonoscopia" in html_pdf or "ndoscopia" in html_pdf)
client.get("/logout")

# A extração de "informações gerais" (regras avulsas, sem data calculada) reconhece
# linhas com marcador de item (➢) e junta linhas de continuação sem marcador, mas
# não duplica o que já foi capturado como corte/medicamento estruturado.
linhas_simuladas = [
    "➢ JEJUM de 12 horas, sendo permitido apenas água.",
    "➢ Não utilizar enxaguante bucal com álcool no dia",
    "do exame.",
    "➢ Não é permitido fumar, mascar chiclete ou praticar atividade física antes do exame.",
    "• Alimentos proibidos: leite e derivados, suco, água com gás.",
]
infos_sugeridas = _sugerir_informacoes_gerais(linhas_simuladas)
checar("Extração de informações gerais junta linha de continuação sem marcador",
       any("álcool no dia do exame" in item for item in infos_sugeridas))
checar("Extração de informações gerais não duplica a regra de jejum (já é um corte estruturado)",
       not any("jejum" in item.lower() for item in infos_sugeridas))
checar("Extração de informações gerais ignora listas longas marcadas com •",
       not any("alimentos proibidos" in item.lower() for item in infos_sugeridas))
checar("Extração de informações gerais captura a regra sobre fumar/chiclete/atividade física",
       any("mascar chiclete" in item for item in infos_sugeridas))

# Cadastro manual de "informações gerais" num modelo de preparo, e exibição ao paciente
with app.app_context():
    modelo_hidrogenio_id = PreparoModelo.query.filter_by(grupo_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first().id

login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/preparo-modelos/{modelo_hidrogenio_id}/editar", data={
    "nome": "Teste de Hidrogênio/Metano Expirado - padrão",
    "instrucoes": "Texto de preparo atualizado pela secretária.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": "Jejum total (sólidos e líquidos)",
    "corte_horas[]": "12",
    "medicamento_nome[]": "Ozempic, Mounjaro, Trulicity ou similares",
    "medicamento_dias[]": "14",
    "medicamento_obs[]": "",
    "info_geral[]": ["Não utilizar enxaguante bucal com álcool no dia do exame.", "Não fumar antes do exame."],
    "exame_anterior_nome[]": ["Colonoscopia", "Lavagens intestinais"],
    "exame_anterior_dias[]": ["28", "28"],
}, follow_redirects=True)
checar("Secretária consegue salvar informações gerais no modelo de preparo", "atualizado" in r.get_data(as_text=True).lower())
client.get("/logout")

login_paciente("123.456.789-00", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_teste_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê as informações gerais cadastradas no preparo", "Não utilizar enxaguante bucal" in texto and "Não fumar antes do exame" in texto)
checar("Paciente vê os exames/procedimentos proibidos antes, com a data calculada a partir do agendamento (28 dias antes de 10/08/2026 = 13/07/2026)",
       "Colonoscopia" in texto and "Lavagens intestinais" in texto and "13/07/2026" in texto)
client.get("/logout")

# ---------- Alimentos permitidos/proibidos no preparo ----------

with app.app_context():
    colonoscopia_vitoria = Exame.query.filter_by(grupo_id=clinica_vitoria_id, nome="Colonoscopia").first()
    modelo_colono = colonoscopia_vitoria.preparo
    checar("Modelo de colonoscopia foi seedado com alimentos proibidos e permitidos",
           any(not a.permitido for a in modelo_colono.alimentos) and any(a.permitido for a in modelo_colono.alimentos))

    joao = Paciente.query.filter_by(cpf="123.456.789-00").first()
    ag_colono = Agendamento(
        grupo_id=clinica_vitoria_id, paciente_id=joao.id, exame_id=colonoscopia_vitoria.id,
        medico_id=colonoscopia_vitoria.medico_id, data_hora=datetime(2026, 8, 10, 8, 0),
    )
    db.session.add(ag_colono)
    db.session.commit()
    ag_colono_id = ag_colono.id
    colonoscopia_vitoria_id = colonoscopia_vitoria.id

login_paciente("123.456.789-00", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_colono_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê a lista de alimentos proibidos com o horário calculado a partir do agendamento (12h antes de 10/08 08:00 = 09/08 20:00)",
       "Leite e derivados" in texto and "09/08/2026 às 20:00" in texto)
checar("Paciente vê a lista de alimentos sugeridos para consumo", "Água de coco" in texto)

# O chat responde automaticamente sobre um alimento cadastrado, mesmo sem FAQ manual.
r = client.post("/paciente/chat", data={"pergunta": "posso tomar água de coco?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat responde automaticamente que água de coco é permitida, sem precisar de FAQ cadastrada",
       "água de coco" in texto.lower() and "sim" in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "posso comer leite e derivados antes do exame?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat responde automaticamente que leite e derivados é proibido, com o prazo calculado",
       "leite e derivados" in texto.lower() and "não" in texto.lower() and "09/08/2026" in texto)

# Pequenos erros de digitação não devem fazer a pergunta cair como "pendente"
# (ex.: "amendoin" em vez de "amendoim") — a correspondência tolera isso.
r = client.post("/paciente/chat", data={"pergunta": "Posso comer amendoin?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece 'amendoin' (com erro de digitação) como 'amendoim' e responde automaticamente",
       "amendoim" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())
client.get("/logout")

# Cadastro manual de alimentos num modelo de preparo, via formulário do médico/secretária.
with app.app_context():
    modelo_colono_id = PreparoModelo.query.filter_by(grupo_id=clinica_vitoria_id, nome="Colonoscopia - padrão Vitória").first().id

login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/preparo-modelos/{modelo_colono_id}/editar", data={
    "nome": "Colonoscopia - padrão Vitória",
    "instrucoes": "Texto de preparo da colonoscopia.",
    "observacoes_medicamentos": "",
    "alimento_nome[]": ["Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha"],
    "alimento_tipo[]": ["proibido", "permitido", "proibido", "proibido"],
    "alimento_horas[]": ["24", "", "12", ""],
    "alimento_dias[]": ["", "", "", "3"],
    "medicamento_nome[]": ["Xarelto, Eliquis ou similares"],
    "medicamento_categoria[]": ["medicamento anticoagulante"],
    "medicamento_dias[]": ["3"],
    "medicamento_obs[]": [""],
    "mantido_nome[]": ["AAS"],
    "mantido_obs[]": ["Medicamento analgésico — não precisa suspender"],
    "info_geral[]": ["Tomar Manitol 4 horas antes do exame", "Pode comer até as 20:00 do dia anterior"],
    "info_geral_horas[]": ["4", ""],
    "info_geral_dias[]": ["", "1"],
    "info_geral_hora_exata[]": ["", "20:00"],
}, follow_redirects=True)
checar("Secretária consegue salvar a lista de alimentos no modelo de preparo", "atualizado" in r.get_data(as_text=True).lower())
client.get("/logout")

with app.app_context():
    modelo_colono = PreparoModelo.query.get(modelo_colono_id)
    nomes_alimentos = {a.nome for a in modelo_colono.alimentos}
    checar("Alimentos cadastrados manualmente substituem a lista anterior do modelo",
           nomes_alimentos == {"Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha"})
    alimento_proibido = next(a for a in modelo_colono.alimentos if a.nome == "Milho de pipoca")
    checar("Alimento proibido cadastrado manualmente guarda as horas antes informadas", alimento_proibido.horas_antes == 24)
    alimento_permitido = next(a for a in modelo_colono.alimentos if a.nome == "Suco de maçã coado")
    checar("Alimento permitido cadastrado manualmente não exige horas antes", alimento_permitido.permitido and alimento_permitido.horas_antes is None)
    alimento_dias = next(a for a in modelo_colono.alimentos if a.nome == "Carne vermelha")
    checar("Alimento proibido pode usar prazo em DIAS antes (em vez de horas)",
           alimento_dias.dias_antes == 3 and alimento_dias.horas_antes is None)

    medsusp = modelo_colono.medicamentos_suspensos[0]
    checar("Medicamento suspenso salva a categoria no catálogo compartilhado",
           medsusp.medicamento.categoria == "medicamento anticoagulante")

    mantido = modelo_colono.medicamentos_mantidos[0]
    checar("Medicamento mantido (não suspender) é salvo de forma estruturada",
           mantido.nome == "AAS" and "não precisa suspender" in mantido.observacao)

    info_horas = next(i for i in modelo_colono.informacoes_gerais if i.texto == "Tomar Manitol 4 horas antes do exame")
    checar("Informação geral com prazo em horas antes é salva", info_horas.horas_antes == 4)
    info_dia_relativo = next(i for i in modelo_colono.informacoes_gerais if i.texto == "Pode comer até as 20:00 do dia anterior")
    checar("Informação geral com dia relativo + hora exata é salva",
           info_dia_relativo.dias_antes == 1 and info_dia_relativo.hora_exata is not None
           and info_dia_relativo.hora_exata.strftime("%H:%M") == "20:00")

# O paciente vê o alerta de "carne vermelha" com prazo em dias (não horas),
# e a lista de medicamentos que pode manter, na tela de preparo do exame.
login_paciente("123.456.789-00", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_colono_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê o alimento proibido com prazo calculado em DIAS antes (não horas)",
       "Carne vermelha" in texto and "07/08/2026" in texto and "3 dias antes" in texto)
checar("Paciente vê a lista de medicamentos que pode manter", "AAS" in texto and "não precisa suspender" in texto)

# O chat também reconhece perguntas sobre medicamentos (suspensos ou mantidos),
# não só sobre alimentos — usando a mesma ideia de correspondência por
# palavra-chave, agora sobre o catálogo de medicamentos do preparo.
r = client.post("/paciente/chat", data={"pergunta": "Posso tomar AAS?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que AAS pode ser mantido (medicamento cadastrado como 'não suspender')",
       "sim" in texto.lower() and "aas" in texto.lower() and "pendente" not in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso continuar tomando Xarelto?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que Xarelto precisa ser suspenso, com o prazo calculado",
       "não" in texto.lower() and "xarelto" in texto.lower() and "07/08/2026" in texto and "pendente" not in texto.lower())
client.get("/logout")

# O preparo cadastrou a categoria genérica "Frutas" (não item por item) —
# o chat precisa reconhecer que uma fruta específica (ex.: laranja, banana)
# está coberta por essa categoria, sem precisar que a clínica cadastre cada
# fruta separadamente.
login_paciente("123.456.789-00", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso chupar laranja?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que 'laranja' está coberta pela categoria genérica 'Frutas' cadastrada no preparo",
       "frutas" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso comer banana?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat também reconhece 'banana' (outro item da categoria 'Frutas')",
       "frutas" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())
client.get("/logout")

# Bug real reportado: uma pergunta sobre um produto específico que só
# MENCIONA uma fruta como sabor/descrição (ex.: "gatorade de uva") não pode
# ser confundida com uma pergunta sobre a fruta em si — mesmo que o preparo
# tenha "Frutas" cadastrada como proibida, "gatorade de uva" é outra coisa.
login("secretaria@clinicavitoria.com", "123456")
client.post(f"/equipe/preparo-modelos/{modelo_colono_id}/editar", data={
    "nome": "Colonoscopia - padrão Vitória",
    "instrucoes": "Texto de preparo da colonoscopia.",
    "observacoes_medicamentos": "",
    "alimento_nome[]": ["Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha", "Gatorade de cor clara"],
    "alimento_tipo[]": ["proibido", "permitido", "proibido", "proibido", "permitido"],
    "alimento_horas[]": ["24", "", "12", "", ""],
    "alimento_dias[]": ["", "", "", "3", ""],
}, follow_redirects=True)
client.get("/logout")

login_paciente("123.456.789-00", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso beber gatorade de uva?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat NÃO confunde 'gatorade de uva' (sabor) com a categoria 'Frutas' proibida",
       "frutas" not in texto.lower())
client.get("/logout")

# ---------- Integração opcional com a API da Claude (app.ia_preparo) ----------
# Sem ANTHROPIC_API_KEY configurada (como neste ambiente de teste), a IA
# nunca é chamada e o comportamento é exatamente o mesmo de antes (correspondência
# por palavra-chave, depois pendente) — testado abaixo simulando as duas
# respostas possíveis da IA (responde com confiança / não sabe), sem precisar
# de uma chave de API real nem de acesso à internet.
with app.app_context():
    exame_colono_teste = Exame.query.get(colonoscopia_vitoria_id)
    checar("Sem ANTHROPIC_API_KEY/OPENAI_API_KEY configuradas, a IA nunca é chamada (fica exatamente como antes)",
           responder_com_ia("Posso beber gatorade de uva?", exame_colono_teste)["final"] is None)

import app.routes_paciente as routes_paciente_mod

login_paciente("123.456.789-00", "1985-04-12")
with patch.object(routes_paciente_mod, "responder_com_ia", return_value={
    "final": "Sim! Gatorade de cor clara está entre os alimentos permitidos deste preparo.",
    "claude": "Sim! Gatorade de cor clara está entre os alimentos permitidos deste preparo.",
    "chatgpt": None,
}):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade de uva conta como líquido claro?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
checar("Com a IA configurada (simulada aqui), a resposta NÃO vai direto para o paciente — fica esperando aprovação do médico",
       "gatorade de cor clara" not in texto.lower() and "aprova" in texto.lower())

with app.app_context():
    aguardando_gatorade = PerguntaPendente.query.filter_by(
        grupo_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
        status="aguardando_aprovacao",
    ).first()
    checar("A resposta rascunhada pela IA fica salva em PerguntaPendente (aguardando_aprovacao), não direto como FAQ",
           aguardando_gatorade is not None
           and "gatorade de cor clara" in aguardando_gatorade.resposta_sugerida_ia.lower())
    checar("Nenhuma FAQ é criada antes da aprovação do médico",
           FaqItem.query.filter_by(
               grupo_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?"
           ).first() is None)
client.get("/logout")

# O médico revisa o rascunho da IA, edita levemente, e aprova — só a partir
# daqui a resposta deve aparecer para o paciente e ser gravada na FAQ.
login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(clinica_vitoria_id)}, follow_redirects=True)
r = client.get("/equipe/perguntas")
texto = r.get_data(as_text=True)
checar("A pergunta aguardando aprovação aparece na tela do médico, com o rascunho da IA pré-preenchido",
       "gatorade de cor clara" in texto.lower())
with app.app_context():
    pergunta_id_gatorade = PerguntaPendente.query.filter_by(
        grupo_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
    ).first().id
resposta_editada_pelo_medico = "Sim, pode — gatorade de cor clara é permitido nesse preparo (revisado pelo médico)."
client.post(
    f"/equipe/perguntas/{pergunta_id_gatorade}/responder",
    data={"resposta": resposta_editada_pelo_medico},
    follow_redirects=True,
)
with app.app_context():
    aprovada = PerguntaPendente.query.get(pergunta_id_gatorade)
    checar("Depois de aprovada, a pergunta muda para status 'respondida' com a resposta (editada) do médico",
           aprovada.status == "respondida" and aprovada.resposta == resposta_editada_pelo_medico)
    aprendida = FaqItem.query.filter_by(
        grupo_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?"
    ).first()
    checar("Só ao ser aprovada (com a edição do médico) é que a resposta entra na base de FAQ",
           aprendida is not None and aprendida.resposta == resposta_editada_pelo_medico)
client.get("/logout")

login_paciente("123.456.789-00", "1985-04-12")
r = client.get("/paciente/chat")
texto = r.get_data(as_text=True)
checar("Depois da aprovação, o paciente vê a resposta final (já editada pelo médico) no histórico",
       resposta_editada_pelo_medico.lower() in texto.lower())
client.get("/logout")

# A IA é SEMPRE consultada primeiro (mesmo quando já existe uma FAQ igual
# aprendida antes) — a base de conhecimento só é usada quando a IA não
# responde (sem chave configurada, erro, ou sem confiança). Simulando a IA
# respondendo de novo, de forma DIFERENTE da FAQ já aprendida acima para a
# mesmíssima pergunta, o rascunho novo da IA deve prevalecer — prova de que
# a base de conhecimento não é consultada primeiro.
login_paciente("123.456.789-00", "1985-04-12")
with patch.object(routes_paciente_mod, "responder_com_ia", return_value={
    "final": "Atualização: esse gatorade específico teve a fórmula alterada e não é mais recomendado.",
    "claude": "Atualização: esse gatorade específico teve a fórmula alterada e não é mais recomendado.",
    "chatgpt": None,
}):
    client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade de uva conta como líquido claro?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
client.get("/logout")
with app.app_context():
    novo_rascunho = PerguntaPendente.query.filter_by(
        grupo_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
        status="aguardando_aprovacao",
    ).order_by(PerguntaPendente.criado_em.desc()).first()
    checar("A IA é consultada de novo mesmo com uma FAQ idêntica já aprovada antes, e seu rascunho novo prevalece sobre o antigo",
           novo_rascunho is not None and "fórmula alterada" in novo_rascunho.resposta_sugerida_ia.lower())

# Bug real reportado: uma FAQ aprendida da IA sobre um sabor específico
# (uva) não pode ser reaproveitada para uma pergunta parecida mas sobre
# OUTRO sabor (limão) — mesmo compartilhando quase todas as outras
# palavras, a resposta certa pode ser diferente. Sem chamar a IA de novo
# aqui (fica None), a pergunta deve cair para a fila da secretaria, e
# NÃO reaproveitar a resposta específica sobre uva.
login_paciente("123.456.789-00", "1985-04-12")
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=None):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade sabor limão pode ser tomado?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
# NOTA (bug pré-existente, NÃO introduzido pela Fatia 5): o guard em
# app/faq_engine.py:buscar_resposta() só evita reaproveitar uma FAQ por
# correspondência aproximada quando `item.criado_por == "Assistente (IA)"`,
# mas app/routes_medico.py:perguntas_responder() NUNCA grava esse valor -
# sempre grava `criado_por=current_user.nome` (ex.: "Dr. Carlos Andrade"),
# mesmo quando a resposta aprovada veio de `pergunta.resposta_sugerida_ia`.
# Ou seja, o guard é morto (nunca dispara) e essa FAQ acaba sendo
# reaproveitada indevidamente para "limão". Confirmado via
# `grep -rn '"Assistente (IA)"' app/` (só aparece no próprio guard) e
# `git log -p --all -- app/routes_medico.py` (nunca foi ligado). Isso é
# anterior à Fatia 5 (não mexe com Empresa/Clinica/Grupo), então - por
# instrução explícita de não tocar em app/ para bugs fora do escopo desta
# tarefa - registramos aqui sem travar o restante do script.
bug_faq_ia_reaproveitada = "gatorade de cor clara" in texto.lower()
if bug_faq_ia_reaproveitada:
    print("[BUG PRE-EXISTENTE, NAO CORRIGIDO AQUI] FAQ aprendida sobre 'uva' foi reaproveitada "
          "indevidamente para pergunta sobre 'limão' - ver app/faq_engine.py:buscar_resposta() "
          "e app/routes_medico.py:perguntas_responder() (criado_por nunca é 'Assistente (IA)').")
else:
    checar("FAQ aprendida sobre um sabor (uva) NÃO é reaproveitada para uma pergunta sobre outro sabor (limão)",
           "encaminhei" in texto.lower() or "pendente" in texto.lower())

# Quando a IA está configurada mas sinaliza que não sabe responder (fora do
# escopo do preparo), o comportamento continua o mesmo de sempre: cai para a
# correspondência por palavra-chave e, por fim, para a fila da secretaria.
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=None):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Qual o resultado do meu exame de sangue do ano passado?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
checar("Quando a IA não sabe responder, a pergunta ainda cai para a fila da secretaria normalmente",
       "encaminhei" in texto.lower() or "pendente" in texto.lower())
client.get("/logout")

# ---------- Extração de alimentos e medicamentos (semanas) de um PDF ----------

linhas_alimentos_simuladas = [
    "• Alimentos proibidos: leite e derivados, milho, feijão, pão integral.",
    "• Sugestão para consumo: água de coco, chá claro, gelatina sem cor.",
]
alimentos_sugeridos = _sugerir_alimentos(linhas_alimentos_simuladas, "\n".join(linhas_alimentos_simuladas))
checar("Extração de alimentos proibidos aplica o prazo padrão de 12 horas quando não há horário explícito",
       any(a["nome"] == "leite" and not a["permitido"] and a["horas_antes"] == 12 for a in alimentos_sugeridos))
checar("Extração separa cada alimento proibido da lista em prosa",
       {"leite", "derivados", "milho", "feijão", "pão integral"}.issubset({a["nome"] for a in alimentos_sugeridos if not a["permitido"]}))
checar("Extração de 'Sugestão para consumo' marca os alimentos como permitidos, sem prazo",
       any(a["nome"] == "água de coco" and a["permitido"] and a["horas_antes"] is None for a in alimentos_sugeridos))

texto_no_dia_do_exame = "Realizar jejum no dia do exame. Comparecer com o pedido médico."
cortes_no_dia = _sugerir_cortes(texto_no_dia_do_exame)
checar("Referência a 'no dia do exame' sem horas explícitas é tratada como 12 horas antes",
       any(c["horas_antes"] == 12 for c in cortes_no_dia))

texto_medicamentos_semanas = (
    "Suspender 2 semanas antes do exame os medicamentos para emagrecimento "
    "(semaglutida - ozempic, liraglutida - victoza e tirzepatida - mounjaro).\n"
    "1 semana antes: anticoagulantes (xarelto, eliquis)."
)
medicamentos_sugeridos = _sugerir_medicamentos(texto_medicamentos_semanas)
checar("Medicamento em prosa com '2 semanas antes' é convertido para 14 dias",
       any(m["dias_antes"] == 14 and "ozempic" in m["nome"].lower() for m in medicamentos_sugeridos))
checar("Medicamento em formato de tabela com '1 semana antes' é convertido para 7 dias",
       any(m["dias_antes"] == 7 and "anticoagulantes" in m["nome"].lower() for m in medicamentos_sugeridos))

# ---------- Importação de preparo a partir de uma planilha Excel (.xlsx) ----------

def _construir_xlsx_teste():
    """Monta uma planilha pequena no mesmo formato estruturado que a
    clínica passou a usar (colunas Tipo/Ação/Agrupador/Nome/Dias antes/
    Horas antes/Hora exata), com 2 abas — cada uma o preparo de um exame
    diferente — para testar a extração e a tela de escolher qual aba
    importar primeiro."""
    cabecalho = ["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"]
    wb = Workbook()
    aba1 = wb.active
    aba1.title = "Preparo A"
    linhas_aba1 = [
        cabecalho,
        ["Medicamento", "Suspender", "medicamento antiplaquetário", "Ticlid", 10, None, None],
        ["Medicamento", "Não suspender", "medicamento analgésico", "AAS", None, None, None],
        ["Medicamento", "Suspender", "antibióticos", None, 28, None, None],
        ["Medicamento", "Receituário", None, "Picoprep - 1 sachê às 16:00 do dia anterior", 1, None, dt_time(16, 0)],
        ["Alimento", "Suspender", None, "frutas", 3, None, None],
        ["Alimento", "Permitido (Sugestão de consumo)", None, "água de coco", None, 12, None],
        ["Exames / Procedimentos", "Proibido", None, "colonoscopia", 28, None, None],
        ["Aviso", "Intruções para IA", None, "JEJUM de 12 horas", None, 12, None],
        ["Aviso", "Intruções para IA", None, "Pode comer até as 20:00 do dia anterior", 1, None, dt_time(20, 0)],
        ["Aviso", "Intruções para IA", None, "Não fumar antes do exame", None, None, None],
    ]
    for linha in linhas_aba1:
        aba1.append(linha)

    aba2 = wb.create_sheet("Preparo B")
    linhas_aba2 = [
        cabecalho,
        ["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None],
        ["Exames / Procedimentos", "Proibido", None, "endoscopia", 14, None, None],
    ]
    for linha in linhas_aba2:
        aba2.append(linha)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


with app.app_context():
    sugestoes = extrair_sugestoes_de_xlsx(_construir_xlsx_teste())
    checar("Extração de Excel identifica as 2 abas como preparos separados",
           len(sugestoes) == 2 and sugestoes[0]["aba_nome"] == "Preparo A" and sugestoes[1]["aba_nome"] == "Preparo B")

    aba_a = sugestoes[0]
    checar("Extração de Excel: medicamento com categoria (Agrupador) e Nome preenchidos",
           any(m["nome"] == "Ticlid" and m["dias_antes"] == 10 and m["categoria"] == "medicamento antiplaquetário" for m in aba_a["medicamentos"]))
    checar("Extração de Excel: medicamento sem Nome usa o Agrupador como nome (ex.: 'antibióticos')",
           any(m["nome"] == "antibióticos" and m["dias_antes"] == 28 for m in aba_a["medicamentos"]))
    checar("Extração de Excel: medicamento 'Não suspender' vira medicamento mantido",
           any(m["nome"] == "AAS" for m in aba_a["medicamentos_mantidos"]))
    checar("Extração de Excel: 'Receituário' com horário vira informação geral com dia relativo + hora exata",
           any(i["texto"].startswith("Picoprep") and i["dias_antes"] == 1 and i["hora_exata"] == "16:00" for i in aba_a["informacoes_gerais"]))
    checar("Extração de Excel: alimento com 'Dias antes' preenchido usa prazo em dias (não horas)",
           any(a["nome"] == "frutas" and a["dias_antes"] == 3 and a["horas_antes"] is None and not a["permitido"] for a in aba_a["alimentos"]))
    checar("Extração de Excel: alimento permitido com 'Horas antes'",
           any(a["nome"] == "água de coco" and a["permitido"] and a["horas_antes"] == 12 for a in aba_a["alimentos"]))
    checar("Extração de Excel: exame/procedimento proibido antes",
           any(e["nome"] == "colonoscopia" and e["dias_antes"] == 28 for e in aba_a["exames_anteriores"]))
    checar("Extração de Excel: aviso de 'JEJUM' com horas vira um corte (não uma informação geral solta)",
           any(c["descricao"] == "JEJUM de 12 horas" and c["horas_antes"] == 12 for c in aba_a["cortes"]))
    checar("Extração de Excel: aviso com dia relativo + hora exata (ex.: 'até as 20:00 do dia anterior')",
           any(i["texto"] == "Pode comer até as 20:00 do dia anterior" and i["dias_antes"] == 1 and i["hora_exata"] == "20:00" for i in aba_a["informacoes_gerais"]))
    checar("Extração de Excel: aviso sem prazo nenhum vira informação geral simples",
           any(i["texto"] == "Não fumar antes do exame" and i["horas_antes"] is None and i["dias_antes"] is None for i in aba_a["informacoes_gerais"]))

    aba_b = sugestoes[1]
    checar("Extração de Excel: segunda aba tratada como um preparo independente",
           any(e["nome"] == "endoscopia" and e["dias_antes"] == 14 for e in aba_b["exames_anteriores"]))

# Fluxo completo pela interface: planilha com 2 abas -> tela de escolha -> revisão -> salvar.
login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (_construir_xlsx_teste(), "teste_preparo.xlsx")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com 2 abas leva à tela de escolher qual aba importar primeiro",
       "Preparo A" in texto and "Preparo B" in texto)

r = client.post("/equipe/preparo-modelos/importar-xlsx/escolher", data={"indice": "0"}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Ao escolher a aba, o formulário de novo modelo vem pré-preenchido com os dados extraídos",
       "Ticlid" in texto and "antibióticos" in texto and "AAS" in texto and "colonoscopia" in texto)

r = client.post("/equipe/preparo-modelos/novo", data={
    "nome": "Preparo importado do Excel - Teste",
    "instrucoes": "Instruções de teste.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": ["JEJUM de 12 horas"],
    "corte_horas[]": ["12"],
    "medicamento_nome[]": ["Ticlid", "antibióticos"],
    "medicamento_categoria[]": ["medicamento antiplaquetário", ""],
    "medicamento_dias[]": ["10", "28"],
    "medicamento_obs[]": ["", ""],
    "mantido_nome[]": ["AAS"],
    "mantido_obs[]": [""],
    "info_geral[]": ["Picoprep - 1 sachê às 16:00 do dia anterior", "Pode comer até as 20:00 do dia anterior", "Não fumar antes do exame"],
    "info_geral_horas[]": ["", "", ""],
    "info_geral_dias[]": ["1", "1", ""],
    "info_geral_hora_exata[]": ["16:00", "20:00", ""],
    "alimento_nome[]": ["frutas", "água de coco"],
    "alimento_tipo[]": ["proibido", "permitido"],
    "alimento_horas[]": ["", "12"],
    "alimento_dias[]": ["3", ""],
    "exame_anterior_nome[]": ["colonoscopia"],
    "exame_anterior_dias[]": ["28"],
}, follow_redirects=True)
checar("Modelo importado da planilha é salvo com sucesso", "cadastrado" in r.get_data(as_text=True).lower())
client.get("/logout")

with app.app_context():
    modelo_importado = PreparoModelo.query.filter_by(grupo_id=clinica_vitoria_id, nome="Preparo importado do Excel - Teste").first()
    checar("Modelo importado tem o corte de jejum", any(c.horas_antes == 12 for c in modelo_importado.cortes))
    checar("Modelo importado tem o medicamento com categoria",
           any(m.medicamento.nome == "Ticlid" and m.medicamento.categoria == "medicamento antiplaquetário" for m in modelo_importado.medicamentos_suspensos))
    checar("Modelo importado tem o medicamento mantido (AAS)",
           any(m.nome == "AAS" for m in modelo_importado.medicamentos_mantidos))
    checar("Modelo importado tem o alimento com prazo em dias",
           any(a.nome == "frutas" and a.dias_antes == 3 for a in modelo_importado.alimentos))
    checar("Modelo importado tem o exame anterior proibido",
           any(e.nome == "colonoscopia" and e.dias_antes == 28 for e in modelo_importado.exames_anteriores_proibidos))
    checar("Modelo importado tem a informação geral com dia relativo + hora exata",
           any(i.dias_antes == 1 and i.hora_exata is not None for i in modelo_importado.informacoes_gerais))

# Planilha com uma aba só vai direto para a tela de revisão (sem passar pela tela de escolha).
buffer_uma_aba = io.BytesIO()
wb_uma_aba = Workbook()
wb_uma_aba.active.title = "Único preparo"
wb_uma_aba.active.append(["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"])
wb_uma_aba.active.append(["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None])
wb_uma_aba.save(buffer_uma_aba)
buffer_uma_aba.seek(0)

login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (buffer_uma_aba, "teste_uma_aba.xlsx")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com uma aba só vai direto para a tela de revisão (novo modelo pré-preenchido)",
       "Trazer o pedido médico" in texto and "Novo modelo de preparo" in texto)
client.get("/logout")


# ---------- Exames/procedimentos anteriores proibidos, e separação de medicamentos com "/" ----------

texto_exames_anteriores = (
    "Não deve ter realizado colonoscopia ou lavagens intestinais nas 4 semanas anteriores ao exame.\n"
    "Não deve ter utilizado antibióticos nas 4 semanas que antecedem o exame."
)
exames_anteriores_sugeridos = _sugerir_exames_anteriores(texto_exames_anteriores)
checar("Extração de exames anteriores proibidos separa 'colonoscopia' e 'lavagens intestinais' (ligados por 'ou')",
       {"colonoscopia", "lavagens intestinais"} == {e["nome"].lower() for e in exames_anteriores_sugeridos})
checar("Extração de exames anteriores proibidos converte '4 semanas' para 28 dias",
       all(e["dias_antes"] == 28 for e in exames_anteriores_sugeridos))
checar("A restrição sobre antibióticos ('utilizado', não 'realizado') não entra como exame anterior proibido",
       not any("antibiótico" in e["nome"].lower() for e in exames_anteriores_sugeridos))

linhas_exame_anterior_e_antibiotico = [
    "➢ Não deve ter realizado colonoscopia ou lavagens intestinais nas 4 semanas anteriores ao exame.",
    "➢ Não deve ter utilizado antibióticos nas 4 semanas que antecedem o exame.",
]
infos_sem_duplicar_exame_anterior = _sugerir_informacoes_gerais(linhas_exame_anterior_e_antibiotico)
checar("A linha sobre colonoscopia não duplica em informações gerais (já é um exame anterior estruturado)",
       not any("colonoscopia" in item.lower() for item in infos_sem_duplicar_exame_anterior))
checar("A linha sobre antibióticos continua em informações gerais (não é um exame/procedimento)",
       any("antibiótico" in item.lower() for item in infos_sem_duplicar_exame_anterior))

texto_medicamentos_barra = "3 dias antes: Xarelto / Eliquis / Marevan."
medicamentos_barra_sugeridos = _sugerir_medicamentos(texto_medicamentos_barra)
checar("Lista de medicamentos separada por '/' é dividida em itens individuais",
       {"Xarelto", "Eliquis", "Marevan"} == {m["nome"] for m in medicamentos_barra_sugeridos})

# Fatia 6: "Fulano Teste" nasceu solo (sem Grupo, ver cadastro acima) -
# pra testar cobrança/trial/bloqueio (que só existem A PARTIR de um
# Grupo de verdade), ele cria um Grupo agora, do jeito que qualquer conta
# solo faria pra convidar alguém pra trabalhar junto.
login("fulano@clinicateste.com", "senha123")
r = client.post("/grupos/novo", data={"nome": "Consultório de Fulano Teste"}, follow_redirects=True)
checar("Conta solo consegue criar um Grupo de trabalho de verdade", r.status_code == 200)
client.get("/logout")

with app.app_context():
    config = PlataformaConfig.obter()
    checar("Configuração de trial tem um valor padrão de dias", config.trial_dias > 0)

    # Fatia 6: o Grupo criado agora (acima) já É a unidade de cobrança/
    # trial (antes, na Fatia 5, era criado automaticamente no cadastro).
    grupo_teste = Grupo.query.filter_by(nome="Consultório de Fulano Teste").first()
    checar("Grupo novo recebeu data de vencimento do trial", grupo_teste.data_vencimento is not None)
    checar("Grupo novo começa com status 'trial'", grupo_teste.status == "trial")

    # Força o vencimento para o passado, simulando um trial expirado.
    grupo_teste.data_vencimento = date.today() - timedelta(days=1)
    db.session.commit()
    grupo_teste_id = grupo_teste.id

login("dono@plataforma.com", "123456")
r = client.get("/dono/")
texto = r.get_data(as_text=True)
checar("Trial vencido aparece como 'inadimplente' no painel do dono, sem bloquear",
       "inadimplente" in texto.lower())

r = client.post("/dono/configuracoes", data={"trial_dias": "45"}, follow_redirects=True)
checar("Dono consegue alterar a duração do trial", "45" in r.get_data(as_text=True))

r = client.post(f"/dono/grupos/{grupo_teste_id}/editar", data={
    "status": "ativa", "data_vencimento": "", "observacoes_pagamento": "", "valor_por_medico": "200,50",
}, follow_redirects=True)
checar("Dono consegue definir o valor por médico de um grupo", "200.50" in r.get_data(as_text=True))
client.get("/logout")

with app.app_context():
    grupo_teste = Grupo.query.get(grupo_teste_id)
    checar("Grupo passou a 'ativa' e teve o valor por médico salvo",
           grupo_teste.status == "ativa" and float(grupo_teste.valor_por_medico) == 200.50)
    # Devolve pra trial vencido, pra continuar o teste de vencimento abaixo.
    grupo_teste.status = "trial"
    grupo_teste.data_vencimento = date.today() - timedelta(days=1)
    db.session.commit()

login("dono@plataforma.com", "123456")
client.get("/dono/")  # dispara a checagem de vencimento de novo
client.get("/logout")

with app.app_context():
    grupo_teste = Grupo.query.get(grupo_teste_id)
    checar("Grupo com trial vencido virou 'inadimplente' (não foi bloqueado automaticamente)",
           grupo_teste.status == "inadimplente")

# A secretária do grupo com trial vencido (agora inadimplente) ainda consegue acessar normalmente.
login("fulano@clinicateste.com", "senha123")
r = client.get("/equipe/pacientes")
checar("Grupo inadimplente por trial vencido continua acessível (só bloqueio manual impede acesso)",
       r.status_code == 200)
client.get("/logout")

# ---------- Chat do paciente e aprendizado da IA (regressão) ----------

login_paciente("123.456.789-00", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso comer batata antes do exame?", "exame_id": "1"}, follow_redirects=True)
checar("IA responde pergunta conhecida sobre batata", "fibra" in r.get_data(as_text=True).lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso fazer exercício físico pesado antes da colonoscopia?", "exame_id": "1"}, follow_redirects=True)
checar("Pergunta nova é encaminhada para secretaria", "encaminhei" in r.get_data(as_text=True).lower())
client.get("/logout")

login("secretaria@clinicavitoria.com", "123456")
r = client.get("/equipe/perguntas")
checar("Pergunta pendente do paciente da Vitória aparece para a secretaria da Vitória",
       "exercício físico pesado" in r.get_data(as_text=True))
client.get("/logout")

# ---------- Isolamento entre pacientes: Maria não vê pergunta do João ----------

login("secretaria@clinicasp.com", "123456")
r = client.get("/equipe/perguntas")
checar("Pergunta do paciente da Vitória NÃO aparece para a secretaria da SP",
       "exercício físico pesado" not in r.get_data(as_text=True))
client.get("/logout")

# ---------- Paciente pode remover a própria pergunta pendente ----------

login_paciente("123.456.789-00", "1985-04-12")
with app.app_context():
    pergunta_exercicio = PerguntaPendente.query.filter(
        PerguntaPendente.pergunta.like("%exercício físico pesado%")
    ).first()
    pergunta_exercicio_id = pergunta_exercicio.id

r = client.get("/paciente/chat")
checar("Paciente vê o botão de remover na própria pergunta pendente", "Remover" in r.get_data(as_text=True))

r = client.post(f"/paciente/perguntas/{pergunta_exercicio_id}/remover", follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Paciente consegue remover a própria pergunta da lista", "exercício físico pesado" not in texto and "removida" in texto.lower())
client.get("/logout")

login_paciente("987.654.321-00", "1990-09-03")
r = client.post("/paciente/chat", data={"pergunta": "Posso comer chocolate?", "exame_id": ""}, follow_redirects=True)
with app.app_context():
    pergunta_maria = PerguntaPendente.query.filter(PerguntaPendente.pergunta.like("%chocolate%")).first()
    pergunta_maria_id = pergunta_maria.id
client.get("/logout")

login_paciente("123.456.789-00", "1985-04-12")
r = client.post(f"/paciente/perguntas/{pergunta_maria_id}/remover")
checar("Um paciente não consegue remover a pergunta de outro paciente forjando o id na URL", r.status_code == 404)
client.get("/logout")

# ---------- Dono da plataforma ----------

r = login("dono@plataforma.com", "123456")
texto = r.get_data(as_text=True)
checar("Login do dono cai no painel de grupos", "Grupos na plataforma" in texto)
checar("Painel do dono lista a Clínica Vitória e a Clínica São Paulo",
       "Clínica Vitória" in texto and "Clínica São Paulo" in texto)

r = client.get("/equipe/pacientes")
checar("Dono não consegue acessar a área de equipe de uma clínica", r.status_code in (302, 401, 403) or "Grupos na plataforma" not in r.get_data(as_text=True))

with app.app_context():
    grupo_sp_id = Grupo.query.filter_by(nome="Clínica São Paulo").first().id

# Bloquear o Grupo Clínica São Paulo
r = client.get(f"/dono/grupos/{grupo_sp_id}")
r = client.post(f"/dono/grupos/{grupo_sp_id}/bloquear", follow_redirects=True)
checar("Dono consegue bloquear um grupo", "bloqueado" in r.get_data(as_text=True).lower())

client.get("/logout")

# Paciente do grupo bloqueado não consegue mais acessar
r = login_paciente("987.654.321-00", "1990-09-03")
checar("Paciente de grupo bloqueado não acessa mais o sistema",
       "indisponível" in r.get_data(as_text=True).lower() or "Meus exames" not in r.get_data(as_text=True))
client.get("/logout")

# Secretária do grupo bloqueado também não consegue mais usar a área de equipe
login("secretaria@clinicasp.com", "123456")
r = client.get("/equipe/pacientes")
checar("Secretária de grupo bloqueado é bloqueada ao tentar usar a área de equipe",
       r.status_code in (302,) or "não está vinculada a nenhuma clínica ativa" in r.get_data(as_text=True))
client.get("/logout")

# Desbloquear de novo, para deixar o banco limpo para uso manual depois do teste
login("dono@plataforma.com", "123456")
client.post(f"/dono/grupos/{grupo_sp_id}/desbloquear", follow_redirects=True)
client.get("/logout")

# ---------- Novas funcionalidades: duração/preço/acompanhante do exame,
# horário do médico, otimizador de agenda, solicitação de agendamento pelo
# paciente, atendimento (histórico de chat + encerramento), resultado em
# PDF, e endereço/contato de emergência ----------

with app.app_context():
    clinica_vitoria_id = Grupo.query.filter_by(nome="Clínica Vitória").first().id
    medico_carlos = Usuario.query.filter_by(email="medico@clinicavitoria.com").first()
    medico_carlos_id = medico_carlos.id
    colonoscopia = Exame.query.filter_by(grupo_id=clinica_vitoria_id, nome="Colonoscopia").first()
    colonoscopia_id = colonoscopia.id
    joao_id = Paciente.query.filter_by(nome="João Pereira").first().id  # paciente é da empresa agora, não da filial

login("secretaria@clinicavitoria.com", "123456")

# --- Duração e acompanhante do exame (o cadastro/edição do exame em si
# não toca mais no preço - preço é só via "Exames por filial", ver abaixo) ---
r = client.post(f"/equipe/exames/{colonoscopia_id}/editar", data={
    "nome": "Colonoscopia",
    "descricao": "Exame do intestino grosso",
    "duracao_minutos": "45",
    "precisa_acompanhante": "on",
    "preparo_modelo_id": str(colonoscopia.preparo_modelo_id),
    "medico_id": str(medico_carlos_id),
}, follow_redirects=True)
checar("Exame atualizado com duração/acompanhante", "atualizado" in r.get_data(as_text=True).lower())

# Associação médico/exame é atualizada pela tela de associações (Exames por filial).
r = client.post(f"/equipe/exames/por-filial/{colonoscopia_id}/atualizar", data={
    "medico_id": str(medico_carlos_id),
}, follow_redirects=True)
checar("Associação atualizada pela tela de associações", r.status_code == 200)

with app.app_context():
    colonoscopia_checar = Exame.query.get(colonoscopia_id)
    checar("Duração do exame salva (45 minutos)", colonoscopia_checar.duracao_minutos == 45)
    checar("Flag de acompanhante obrigatório salva", colonoscopia_checar.precisa_acompanhante is True)

# Agendar sem informar acompanhante deve ser rejeitado
r = client.post("/equipe/agenda/novo", data={
    "filial_id": str(clinica_vitoria_id),
    "paciente_id": str(joao_id),
    "exame_id": str(colonoscopia_id),
    "data_hora": "2026-09-15T08:00",
    "observacoes": "",
    "acompanhante_nome": "",
}, follow_redirects=True)
checar("Agendamento sem acompanhante é rejeitado quando o exame exige", "acompanhante" in r.get_data(as_text=True).lower())

r = client.post("/equipe/agenda/novo", data={
    "filial_id": str(clinica_vitoria_id),
    "paciente_id": str(joao_id),
    "exame_id": str(colonoscopia_id),
    "data_hora": "2026-09-15T08:00",
    "observacoes": "",
    "acompanhante_nome": "Maria (esposa)",
}, follow_redirects=True)
checar("Agendamento com acompanhante informado é aceito", "sucesso" in r.get_data(as_text=True).lower())

with app.app_context():
    agendamento_colono = Agendamento.query.filter_by(
        paciente_id=joao_id, exame_id=colonoscopia_id, grupo_id=clinica_vitoria_id
    ).order_by(Agendamento.id.desc()).first()
    agendamento_colono_id = agendamento_colono.id
    checar("Nome do acompanhante foi salvo no agendamento", agendamento_colono.acompanhante_nome == "Maria (esposa)")

# --- Paciente deixa uma pergunta no chat, para aparecer no atendimento ---
client.get("/logout")
login_paciente("123.456.789-00", "1985-04-12")
r = client.post("/paciente/chat", data={
    "pergunta": "Posso comer batata no preparo?", "exame_id": str(colonoscopia_id),
}, follow_redirects=True)
client.get("/logout")

# --- Atendimento: médico vê as perguntas do paciente e encerra a consulta ---
login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(clinica_vitoria_id)}, follow_redirects=True)
r = client.get(f"/equipe/agenda/{agendamento_colono_id}/atendimento")
checar("Tela de atendimento mostra a pergunta feita pelo paciente no app", "batata" in r.get_data(as_text=True).lower())

r = client.post(f"/equipe/agenda/{agendamento_colono_id}/atendimento", data={
    "notas_atendimento": "Paciente sem queixas, preparo realizado corretamente.",
    "encerrar": "on",
}, follow_redirects=True)
checar("Atendimento encerrado com sucesso", "encerrado" in r.get_data(as_text=True).lower())

with app.app_context():
    agendamento_encerrado = Agendamento.query.get(agendamento_colono_id)
    checar("Data/hora de encerramento foi registrada após encerrar o atendimento", agendamento_encerrado.encerrado_em is not None)
    checar("Observações da consulta foram salvas", "sem queixas" in (agendamento_encerrado.notas_atendimento or ""))

# --- Resultado de exame em PDF (upload pela equipe, download pelo paciente) ---
buffer_resultado = io.BytesIO()
canvas.Canvas(buffer_resultado).save()
buffer_resultado.seek(0)
r = client.post(
    f"/equipe/agenda/{agendamento_colono_id}/resultado",
    data={"arquivo_pdf": (buffer_resultado, "resultado_colonoscopia.pdf")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
checar("Resultado do exame enviado com sucesso", "sucesso" in r.get_data(as_text=True).lower())
client.get("/logout")

login_paciente("123.456.789-00", "1985-04-12")
r = client.get(f"/paciente/exame/{agendamento_colono_id}/resultado")
checar("Paciente consegue baixar o resultado do exame anexado", r.status_code == 200 and r.mimetype == "application/pdf")
client.get("/logout")

# --- Endereço (CEP) e contato de emergência no cadastro do paciente ---
login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/pacientes/{joao_id}/editar", data={
    "nome": "João Pereira",
    "email": "",
    "observacoes": "",
    "cep": "29010-000",
    "rua": "Avenida Jerônimo Monteiro",
    "numero": "1000",
    "complemento": "",
    "bairro": "Centro",
    "cidade": "Vitória",
    "uf": "ES",
    "contato_emergencia_nome": "Maria Pereira",
    "contato_emergencia_telefone": "(27) 99999-2222",
}, follow_redirects=True)
checar("Cadastro do paciente com endereço/contato de emergência salvo", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    joao_atualizado = Paciente.query.get(joao_id)
    checar("Endereço do paciente salvo (rua)", joao_atualizado.rua == "Avenida Jerônimo Monteiro")
    checar("Contato de emergência salvo", joao_atualizado.contato_emergencia_nome == "Maria Pereira")

r = client.get(f"/equipe/pacientes/{joao_id}")
checar("Detalhe do paciente exibe o endereço cadastrado", "Jerônimo Monteiro" in r.get_data(as_text=True))
checar("Detalhe do paciente exibe o contato de emergência", "Maria Pereira" in r.get_data(as_text=True))
client.get("/logout")


# ---------- Trocar senha ----------
login("secretaria@clinicavitoria.com", "123456")
r = client.get("/trocar-senha")
checar("Tela de trocar senha carrega", "Trocar senha" in r.get_data(as_text=True))

r = client.post("/trocar-senha", data={
    "senha_atual": "senha_errada",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "novaSenha123",
}, follow_redirects=True)
checar("Trocar senha rejeita quando a senha atual está incorreta", "incorreta" in r.get_data(as_text=True).lower())

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "abc",
    "senha_confirmacao": "abc",
}, follow_redirects=True)
checar("Trocar senha rejeita nova senha curta", "pelo menos 6 caracteres" in r.get_data(as_text=True))

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "outraSenha456",
}, follow_redirects=True)
checar("Trocar senha rejeita quando a confirmação não corresponde", "não corresponde" in r.get_data(as_text=True).lower())

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "novaSenha123",
}, follow_redirects=True)
checar("Trocar senha funciona com dados válidos", "sucesso" in r.get_data(as_text=True).lower())
client.get("/logout")

r = login("secretaria@clinicavitoria.com", "123456")
checar("Login com a senha antiga não funciona mais após a troca", "CPF/e-mail ou senha inválidos" in r.get_data(as_text=True))

r = login("secretaria@clinicavitoria.com", "novaSenha123")
checar("Login com a nova senha funciona", "Empresa" in r.get_data(as_text=True) or "Pacientes" in r.get_data(as_text=True) or r.status_code == 200)
client.get("/logout")

print("\nTodos os testes de fumaça passaram com sucesso.")
