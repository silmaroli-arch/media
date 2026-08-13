"""Testa a nova tela "Gerar Excel a partir de PDF" (medico.preparo_pdf_para_excel),
que substituiu a antiga importação direta de PDF pro formulário de modelo de
preparo (medico.preparo_modelos_importar_pdf, removida): em vez de preencher
o formulário diretamente com uma extração heurística nem sempre confiável,
a pessoa agora baixa uma planilha Excel (.xlsx) já no formato aceito pela
importação de Excel existente, revisa com calma, e importa ela normalmente
pelo botão "Importar de um Excel" no cadastro de modelo."""
import io

from reportlab.pdfgen import canvas
from openpyxl import load_workbook

from app import create_app

app = create_app()
client = app.test_client()


def checar(nome, condicao):
    status = "OK" if condicao else "FALHOU"
    print(f"[{status}] {nome}")
    assert condicao, nome


def login(email, senha="123456"):
    return client.post("/login", data={"email": email, "senha": senha}, follow_redirects=True)


login("secretaria@gruposaude.com")

# ---------- Menu e tela ----------

r = client.get("/equipe/")
checar("O menu lateral tem o novo item 'Gerar Excel a partir de PDF'", "Gerar Excel a partir de PDF" in r.get_data(as_text=True))

r = client.get("/equipe/preparo-modelos/pdf-para-excel")
checar("Tela responde 200", r.status_code == 200)
html = r.get_data(as_text=True)
checar("Tela tem o campo de arquivo PDF", 'name="arquivo_pdf"' in html)
checar("A tela de novo modelo NÃO tem mais botão de importar PDF (só Excel)",
       "Importar de um PDF" not in client.get("/equipe/preparo-modelos/novo").get_data(as_text=True))

# ---------- Gerar a planilha a partir de um PDF de teste ----------

buffer_pdf = io.BytesIO()
c = canvas.Canvas(buffer_pdf)
c.drawString(50, 800, "PREPARO TESTE GERACAO DE EXCEL")
c.drawString(50, 780, "JEJUM de 10 horas, sendo permitido apenas agua.")
c.drawString(50, 760, "7 dias antes: VARFARINA, XARELTO.")
c.showPage()
c.save()
buffer_pdf.seek(0)

r = client.post(
    "/equipe/preparo-modelos/pdf-para-excel",
    data={"arquivo_pdf": (buffer_pdf, "preparo.pdf")},
    content_type="multipart/form-data",
)
checar("Gerar Excel responde 200", r.status_code == 200)
checar("A resposta é um arquivo pra baixar (attachment)", "attachment" in r.headers.get("Content-Disposition", ""))
checar("O nome do arquivo baixado é o esperado", "preparo-extraido-do-pdf.xlsx" in r.headers.get("Content-Disposition", ""))

planilha = load_workbook(io.BytesIO(r.data))
aba = planilha.active
checar("A aba usa o nome sugerido do exame extraído do PDF", aba.title == "PREPARO TESTE GERACAO DE EXCEL")
linhas = [tuple(l) for l in aba.iter_rows(min_row=2, values_only=True)]
checar("A planilha tem a linha do corte de jejum",
       any(l[0] == "Aviso" and "jejum total" in str(l[3]).lower() and l[5] == 10 for l in linhas))
checar("A planilha tem as linhas dos medicamentos com 7 dias de prazo",
       sum(1 for l in linhas if l[0] == "Medicamento" and l[1] == "Suspender" and l[4] == 7) == 2)

# ---------- Erro: sem arquivo ----------

r_sem_arquivo = client.post("/equipe/preparo-modelos/pdf-para-excel", data={}, content_type="multipart/form-data")
checar("Sem arquivo, mostra aviso e não quebra", "Selecione um arquivo PDF" in r_sem_arquivo.get_data(as_text=True))

# ---------- Erro: PDF inválido/corrompido ----------

r_invalido = client.post(
    "/equipe/preparo-modelos/pdf-para-excel",
    data={"arquivo_pdf": (io.BytesIO(b"isso nao e um pdf de verdade"), "invalido.pdf")},
    content_type="multipart/form-data",
)
checar(
    "PDF inválido mostra aviso amigável, não quebra",
    "Não foi possível ler esse PDF" in r_invalido.get_data(as_text=True),
)

client.get("/logout")
print("\nTodos os testes de 'Gerar Excel a partir de PDF' passaram.")
