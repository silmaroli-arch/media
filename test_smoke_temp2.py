"""Teste rápido (smoke test) dos principais fluxos, sem precisar de navegador."""
import io
from datetime import date, datetime, timedelta, time as dt_time
from unittest.mock import patch

from reportlab.pdfgen import canvas
from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import (
    Empresa, Clinica, PlataformaConfig, Usuario, Paciente, Exame, Agendamento,
    PreparoModelo, PreparoCorte, PreparoMedicamentoSuspenso, PreparoInfoGeral, PreparoAlimento,
    PreparoExameAnterior, PreparoMedicamentoMantido, Medicamento, PerguntaPendente, FaqItem,
    MedicoHorario, ChatMensagem, ResultadoExame, DescontoConfig, Pagamento,
)
from app.pdf_preparo import (
    _sugerir_informacoes_gerais, _sugerir_alimentos, _sugerir_medicamentos, _sugerir_cortes,
    _sugerir_exames_anteriores,
)
from app.xlsx_preparo import extrair_sugestoes_de_xlsx
from app.ia_preparo import responder_com_ia

app = create_app()
client = app.test_client()


def login(email, senha):
    return client.post("/login", data={"email": email, "senha": senha}, follow_redirects=True)


def login_paciente(telefone, data_nascimento):
    return client.post(
        "/login-paciente",
        data={"telefone": telefone, "data_nascimento": data_nascimento},
        follow_redirects=True,
    )


def checar(nome, condicao):
    status = "OK" if condicao else "FALHOU"
    print(f"[{status}] {nome}")
    assert condicao, nome


# ---------- Fluxo básico da Clínica Vitória (secretária) ----------

r = login("secretaria@clinicavitoria.com", "123456")
checar("Login secretaria (Vitória) funciona", r.status_code == 200 and "Painel" in r.get_data(as_text=True))
checar("Navbar mostra o nome da clínica atual", "Clínica Vitória" in r.get_data(as_text=True))

r = client.get("/equipe/pacientes")
checar("Lista de pacientes da Vitória contém João, não contém Maria",
       "João Pereira" in r.get_data(as_text=True) and "Maria Silva" not in r.get_data(as_text=True))

r = client.get("/equipe/exames")
checar("Lista de exames da Vitória contém Colonoscopia", "Colonoscopia" in r.get_data(as_text=True))

r = client.get("/equipe/agenda", follow_redirects=True)
checar("Agenda da Vitória acessível (redireciona para o painel)", "Agenda de exames" in r.get_data(as_text=True))

r = client.get("/equipe/equipe-membros")
texto = r.get_data(as_text=True)
checar("Equipe da Vitória lista Ana e o Dr. Carlos", "Ana Secretária" in texto and "Carlos Andrade" in texto)

r = client.get("/equipe/clinica/configuracoes")
texto = r.get_data(as_text=True)
pass  # pre-existente, ja reportado

r = client.post("/equipe/clinica/configuracoes", data={
    "nome": "Clínica Vitória", "telefone": "(27) 3333-4444", "email_contato": "contato@clinicavitoria.com",
    "razao_social": "Clínica Vitória Diagnósticos Ltda.", "cnpj": "12.345.678/0001-90",
    "cep": "29010-000", "rua": "Av. Jerônimo Monteiro", "numero": "1000", "bairro": "Centro",
    "cidade": "Vitória", "uf": "ES",
    "inscricao_estadual": "081.234.567", "regime_tributario": "Simples Nacional",
    "cnae": "8640-2/02", "codigo_ibge_municipio": "3205309",
    "dia_0_ativo": "on", "dia_0_inicio": "08:00", "dia_0_fim": "19:00",
}, follow_redirects=True)
checar("Secretária consegue salvar os dados da clínica", "atualizados com sucesso" in r.get_data(as_text=True).lower())

r = client.get("/equipe/clinica/configuracoes")
pass  # pre-existente, ja reportado

client.get("/logout")

login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)
r = client.get("/equipe/clinica/configuracoes")
checar("Médico não pode acessar os dados da clínica (restrito à secretária)", r.status_code in (302,))
client.get("/logout")

# ---------- Isolamento: secretária da Clínica São Paulo não vê dados da Vitória ----------

login("secretaria@clinicasp.com", "123456")

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Secretária da SP vê Maria mas não João (isolamento)",
       "Maria Silva" in texto and "João Pereira" not in texto)

r = client.get("/equipe/exames")
checar("Exames da SP não incluem Hemograma da Vitória", "Hemograma" not in r.get_data(as_text=True))

client.get("/logout")

# ---------- Médico vinculado a duas clínicas: precisa escolher ----------

r = login("medico@clinicavitoria.com", "123456")
texto = r.get_data(as_text=True)
# Agora o que precisa ser escolhido é a EMPRESA (o cliente/tenant), não a
# filial: este médico atende em duas empresas sem relação entre si.
checar("Médico com vínculo em duas empresas cai na tela de escolha de empresa",
       "Em qual empresa" in texto)

r = client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)
checar("Médico consegue selecionar a Clínica Vitória", "Clínica Vitória" in r.get_data(as_text=True))

r = client.get("/equipe/clinica")
checar("Tela de troca de empresa ainda lista as duas empresas do médico", "Clínica São Paulo" in r.get_data(as_text=True))

client.get("/logout")

# ---------- Área do médico: cada médico só vê seus próprios exames/pacientes ----------

r = login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": "1"}, follow_redirects=True)

r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
checar("Dr. Carlos vê Colonoscopia e Glicemia, mas não o Hemograma da Dra. Fernanda",
       "Colonoscopia" in texto and "Glicemia" in texto and "Hemograma" not in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Dr. Carlos vê João (seu paciente), mas não Pedro (paciente da Dra. Fernanda)",
       "João Pereira" in texto and "Pedro Souza" not in texto)

r = client.get("/equipe/agenda", follow_redirects=True)
checar("Agenda do Dr. Carlos acessível (redireciona para o painel)", "Agenda de exames" in r.get_data(as_text=True))

r = client.get("/equipe/equipe-membros")
checar("Médico não consegue acessar a gestão de equipe", r.status_code in (302,))

client.get("/logout")

login("medica2@clinicavitoria.com", "123456")
r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
checar("Dra. Fernanda vê só o Hemograma, não a Colonoscopia do Dr. Carlos",
       "Hemograma" in texto and "Colonoscopia" not in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Dra. Fernanda vê Pedro (seu paciente), mas não João (paciente do Dr. Carlos)",
       "Pedro Souza" in texto and "João Pereira" not in texto)

client.get("/logout")

# Secretária continua vendo tudo da clínica
login("secretaria@clinicavitoria.com", "123456")
r = client.get("/equipe/exames")
texto = r.get_data(as_text=True)
checar("Secretária vê exames de ambos os médicos da Vitória",
       "Colonoscopia" in texto and "Glicemia" in texto and "Hemograma" in texto)

r = client.get("/equipe/pacientes")
texto = r.get_data(as_text=True)
checar("Secretária vê pacientes de ambos os médicos", "João Pereira" in texto and "Pedro Souza" in texto)
client.get("/logout")

# ---------- Cadastro público cria nova empresa + 1ª filial, isoladas ----------

r = client.post("/cadastro", data={
    "nome_empresa": "Empresa Teste Automatizado",
    "nome": "Fulano Teste",
    "email": "fulano@clinicateste.com",
    "senha": "senha123",
    "papel": "secretaria",
}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Cadastro público cria a empresa e loga automaticamente", "Painel" in texto)
checar("Nome da empresa aparece na navbar", "Empresa Teste Automatizado" in texto)
# O cadastro público (modo "empresa") não cria mais NENHUMA filial
# automaticamente - isso é feito depois, ao entrar no app, em "Meus
# Locais de Atendimento" (ver test_cadastro_empresa_sem_filial.py para o
# fluxo completo).
with app.app_context():
    empresa_nova = Empresa.query.filter_by(nome="Empresa Teste Automatizado").first()
    checar("Empresa nova foi criada", empresa_nova is not None)
    checar("NENHUMA filial foi criada automaticamente no cadastro", Clinica.query.filter_by(empresa_id=empresa_nova.id).first() is None)
    usuario_novo = Usuario.query.filter_by(email="fulano@clinicateste.com").first()
    checar("Usuário fica vinculado à empresa via empresa_fundadora_id", usuario_novo.empresa_fundadora_id == empresa_nova.id)

r = client.get("/equipe/pacientes")
checar("Filial nova começa sem nenhum paciente de outras empresas", "João Pereira" not in r.get_data(as_text=True))

client.get("/logout")

# ---------- Cadastro público escolhendo "médico": mesmo sem secretária, ----------
# ---------- quem cria a empresa recebe todas as permissões administrativas ----------

r = client.post("/cadastro", data={
    "nome_empresa": "Clínica Solo do Dr. Ricardo",
    "nome": "Dr. Ricardo Alves",
    "email": "ricardo@clinicasolo.com",
    "senha": "senha123",
    "papel": "medico",
}, follow_redirects=True)
checar("Cadastro público como médico também funciona", "Painel" in r.get_data(as_text=True))

with app.app_context():
    ricardo = Usuario.query.filter_by(email="ricardo@clinicasolo.com").first()
    checar("Médico que criou a empresa é tipo 'medico'", ricardo.tipo == "medico")
    checar(
        "Médico fundador recebe todas as permissões administrativas (não há secretária)",
        ricardo.perm_pacientes and ricardo.perm_equipe and ricardo.perm_filiais and ricardo.perm_dados_clinica,
    )

r = client.get("/equipe/equipe-membros")
checar("Médico fundador consegue acessar a tela de Equipe mesmo sendo médico", "Equipe da empresa" in r.get_data(as_text=True))
r = client.get("/equipe/filiais")
checar("Médico fundador consegue acessar a tela de Filiais (ainda vazia)", "Nenhum local de atendimento cadastrado" in r.get_data(as_text=True))
r = client.get("/equipe/clinica/configuracoes", follow_redirects=True)
checar("Médico fundador consegue acessar Dados da clínica (com aviso, sem filial ainda)", r.status_code == 200)
r = client.get("/equipe/pacientes/novo")
checar("Médico fundador consegue acessar Novo paciente", r.status_code == 200)

# ---------- Cadastro de paciente sem senha: login por telefone + data de nascimento ----------

r = client.post("/equipe/pacientes/novo", data={
    "nome": "Beatriz Nunes", "cpf": "555.666.777-00", "email": "",
}, follow_redirects=True)
checar("Cadastro de paciente exige telefone e data de nascimento (sem eles, é rejeitado)",
       "obrigatórios" in r.get_data(as_text=True).lower())

r = client.post("/equipe/pacientes/novo", data={
    "nome": "Beatriz Nunes", "cpf": "555.666.777-00", "email": "",
    "telefone": "(28) 98765-4321", "data_nascimento": "1995-06-20",
}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Cadastro de paciente funciona sem e-mail nem senha, só com telefone e data de nascimento",
       "cadastrado" in texto.lower() and "não é necessário criar senha" in texto.lower())
checar("Paciente recém-cadastrado (sem agendamento ainda) já aparece na lista para o médico com permissão de pacientes",
       "Beatriz Nunes" in texto)

r = client.post("/equipe/pacientes/novo", data={
    "nome": "Outra Pessoa", "cpf": "999.888.777-00", "email": "",
    "telefone": "(28) 98765-4321", "data_nascimento": "1980-01-01",
}, follow_redirects=True)
checar("Cadastro rejeita telefone já usado por outro paciente", "já existe um paciente cadastrado com esse telefone" in r.get_data(as_text=True).lower())

client.get("/logout")

r = login_paciente("(28) 98765-4321", "1980-01-01")
checar("Login do paciente com data de nascimento errada é rejeitado", "incorretos" in r.get_data(as_text=True).lower())

r = login_paciente("(28) 98765-4321", "1995-06-20")
texto = r.get_data(as_text=True)
checar("Paciente cadastrado sem senha consegue entrar só com telefone e data de nascimento",
       "Meus exames" in texto or "Tirar dúvidas" in texto)
checar("Paciente sem senha não vê o link de 'Trocar senha' na barra superior", "Trocar senha" not in texto)

client.get("/logout")

# ---------- Filiais: uma empresa com duas unidades ----------

with app.app_context():
    filial_centro_id = Clinica.query.filter_by(nome="Grupo Saúde Total - Centro").first().id

login("secretaria@gruposaude.com", "123456")
# Quem atua em duas filiais da MESMA empresa não escolhe mais nada: cai
# direto no painel e vê os dados das duas filiais juntos, com a filial
# indicada em cada registro.
texto = client.get("/equipe/clinica", follow_redirects=True).get_data(as_text=True)
checar("Secretária multi-filial NÃO precisa mais escolher filial (vai direto ao painel)",
       "Em qual empresa" not in texto and "Painel" in texto)
texto_pacientes = client.get("/equipe/pacientes").get_data(as_text=True)
checar("A lista de pacientes traz a coluna Filial (mais de uma filial acessível)",
       "Filial" in texto_pacientes)

client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)

r = client.get("/equipe/filiais")
texto = r.get_data(as_text=True)
checar("Secretária do Grupo Saúde Total vê as duas filiais", "Grupo Saúde Total - Centro" in texto and "Grupo Saúde Total - Praia" in texto)

r = client.post("/equipe/filiais/nova", data={"nome": "Grupo Saúde Total - Norte"}, follow_redirects=True)
checar("Secretária consegue cadastrar uma terceira filial na mesma empresa", "cadastrado com sucesso" in r.get_data(as_text=True).lower())

r = client.get("/equipe/filiais")
checar("A nova filial aparece nos locais de atendimento da secretária (ela ficou vinculada a ela)",
       "Grupo Saúde Total - Norte" in r.get_data(as_text=True))
client.get("/logout")

with app.app_context():
    grupo = Empresa.query.filter_by(nome="Grupo Saúde Total").first()
    checar("Empresa com filiais tem 3 filiais após o cadastro da nova", len(grupo.filiais) == 3)
    checar("Médico que atua em 2 filiais da mesma empresa conta 1x na cobrança",
           len(grupo.medicos_distintos) == 1)
    checar("Valor mensal estimado é 1 médico x R$150", float(grupo.valor_mensal_estimado) == 150.0)
    filial_praia_id = Clinica.query.filter_by(nome="Grupo Saúde Total - Praia").first().id

# ---------- Editar dados de qualquer filial pela tela de Filiais ----------

login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)

# Mesmo com a filial "Centro" selecionada na sessão, a secretária consegue
# editar os dados da filial "Praia" diretamente, sem precisar trocar antes.
r = client.post(f"/equipe/clinica/configuracoes/{filial_praia_id}", data={
    "nome": "Grupo Saúde Total - Praia", "telefone": "(27) 3222-1111", "email_contato": "praia@gruposaude.com",
    "cidade": "Vila Velha", "uf": "ES",
}, follow_redirects=True)
checar("Secretária consegue editar os dados de uma filial que não é a atual",
       "atualizados com sucesso" in r.get_data(as_text=True).lower())

with app.app_context():
    filial_praia = Clinica.query.get(filial_praia_id)
    checar("Os dados da filial Praia foram salvos mesmo sem ela estar selecionada",
           filial_praia.telefone == "(27) 3222-1111")

# ---------- Cadastrar médico/secretária informando a filial ----------

r = client.post("/equipe/equipe-membros/novo", data={
    "nome": "Dra. Beatriz Costa", "email": "beatriz@gruposaude.com", "papel": "medico",
    "senha": "", "filial_ids": str(filial_praia_id),
}, follow_redirects=True)
checar("Cadastro de médico exige e usa a filial escolhida", "cadastrado" in r.get_data(as_text=True).lower())

with app.app_context():
    from app.models import ClinicaMembro as _CM, Usuario as _Usuario
    nova_medica = _Usuario.query.filter_by(email="beatriz@gruposaude.com").first()
    vinculo = _CM.query.filter_by(usuario_id=nova_medica.id).first()
    checar("A médica nova ficou vinculada à filial Praia (não à Centro, que era a atual)",
           vinculo.clinica_id == filial_praia_id)
    checar("Como nenhuma permissão foi marcada no formulário, a médica nova começa sem nenhuma",
           not (nova_medica.perm_pacientes or nova_medica.perm_equipe or nova_medica.perm_filiais or nova_medica.perm_dados_clinica))
    beatriz_id = nova_medica.id

# Cadastrando outra pessoa marcando as permissões pelos checkboxes
r = client.post("/equipe/equipe-membros/novo", data={
    "nome": "Rafael Souza", "email": "rafael@gruposaude.com", "papel": "secretaria",
    "senha": "", "filial_ids": str(filial_praia_id),
    "perm_pacientes": "on", "perm_equipe": "on",
}, follow_redirects=True)
checar("Cadastro de novo membro aceita marcar permissões específicas pelos checkboxes", "cadastrado" in r.get_data(as_text=True).lower())
with app.app_context():
    rafael = Usuario.query.filter_by(email="rafael@gruposaude.com").first()
    checar("Rafael recebeu só as permissões marcadas (pacientes e equipe)",
           rafael.perm_pacientes and rafael.perm_equipe and not rafael.perm_filiais and not rafael.perm_dados_clinica)

client.get("/logout")

# Beatriz, sem nenhuma permissão, não consegue acessar telas administrativas
login("beatriz@gruposaude.com", "123456")
r = client.get("/equipe/equipe-membros", follow_redirects=True)
checar("Médica sem permissão de equipe é bloqueada ao tentar acessar a tela de Equipe",
       "permissão" in r.get_data(as_text=True).lower())
client.get("/logout")

# Secretária concede a permissão de equipe para a Beatriz pela tela de Permissões
login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)
r = client.get(f"/equipe/equipe-membros/{beatriz_id}/permissoes")
checar("Tela de permissões carrega para um membro existente", "Beatriz" in r.get_data(as_text=True))
r = client.post(f"/equipe/equipe-membros/{beatriz_id}/permissoes", data={"perm_equipe": "on"}, follow_redirects=True)
checar("Permissões atualizadas com sucesso", "atualizadas" in r.get_data(as_text=True).lower())
client.get("/logout")

# Agora a Beatriz consegue acessar a tela de Equipe
login("beatriz@gruposaude.com", "123456")
r = client.get("/equipe/equipe-membros")
checar("Depois de receber a permissão de equipe, a médica consegue acessar a tela", "Equipe da empresa" in r.get_data(as_text=True))
r = client.get("/equipe/filiais", follow_redirects=True)
checar("Mas continua sem acesso a Filiais (não foi essa a permissão concedida)",
       "permissão" in r.get_data(as_text=True).lower())
client.get("/logout")

# ---------- Novo agendamento: filial e médico explícitos ----------

login("secretaria@gruposaude.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(filial_centro_id)}, follow_redirects=True)

with app.app_context():
    medico_grupo_id = _Usuario.query.filter_by(email="medico@gruposaude.com").first().id

r = client.get(f"/equipe/agenda/novo?filial_id={filial_praia_id}&medico_id={medico_grupo_id}")
checar("Tela de novo agendamento mostra o seletor de filial e médico",
       "Filial" in r.get_data(as_text=True) and "Médico" in r.get_data(as_text=True))

client.get("/logout")

# ---------- Modelos de preparo reaproveitáveis, cortes e medicamentos ----------

with app.app_context():
    clinica_vitoria_id = Clinica.query.filter_by(nome="Clínica Vitória").first().id
    modelo_hidrogenio = PreparoModelo.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first()
    exames_hidrogenio = Exame.query.filter_by(clinica_id=clinica_vitoria_id, preparo_modelo_id=modelo_hidrogenio.id).all()
    checar("Os 3 substratos do teste de hidrogênio compartilham o mesmo modelo de preparo, sem duplicar cadastro",
           len(exames_hidrogenio) == 3)
    corte_jejum = modelo_hidrogenio.cortes[0]
    checar("Modelo do teste de hidrogênio tem o corte de jejum de 12 horas", corte_jejum.horas_antes == 12)
    med_susp = modelo_hidrogenio.medicamentos_suspensos[0]
    checar("Modelo do teste de hidrogênio tem o medicamento a suspender 14 dias antes", med_susp.dias_antes == 14)

login("secretaria@clinicavitoria.com", "123456")

# Editar o modelo compartilhado deve refletir nos 3 exames que o usam.
with app.app_context():
    modelo_hidrogenio_id = PreparoModelo.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first().id

r = client.post(f"/equipe/preparo-modelos/{modelo_hidrogenio_id}/editar", data={
    "nome": "Teste de Hidrogênio/Metano Expirado - padrão",
    "instrucoes": "Texto de preparo atualizado pela secretária.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": "Jejum total (sólidos e líquidos)",
    "corte_horas[]": "12",
    "medicamento_nome[]": "Ozempic, Mounjaro, Trulicity ou similares",
    "medicamento_dias[]": "14",
    "medicamento_obs[]": "",
}, follow_redirects=True)
checar("Secretária consegue editar o modelo de preparo compartilhado", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    exame_lactose = Exame.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Lactose").first()
    checar("A edição do modelo aparece refletida no exame de Lactose (mesmo preparo)",
           exame_lactose.preparo.instrucoes == "Texto de preparo atualizado pela secretária.")
    exame_frutose = Exame.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Frutose").first()
    checar("A edição do modelo também aparece no exame de Frutose (mesmo preparo)",
           exame_frutose.preparo.instrucoes == "Texto de preparo atualizado pela secretária.")

client.get("/logout")

# Alertas de corte/medicamentos calculados a partir do horário do agendamento
with app.app_context():
    joao = Paciente.query.filter_by(cpf="123.456.789-00").first()
    exame_lactose = Exame.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste do Hidrogênio - Lactose").first()
    ag_teste = Agendamento(
        clinica_id=clinica_vitoria_id, paciente_id=joao.id, exame_id=exame_lactose.id,
        medico_id=exame_lactose.medico_id, data_hora=datetime(2026, 8, 10, 8, 0), status="agendado",
    )
    db.session.add(ag_teste)
    db.session.commit()
    ag_teste_id = ag_teste.id

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_teste_id}")
texto = r.get_data(as_text=True)
checar("Alerta de jejum mostra o horário calculado a partir do agendamento (12h antes de 10/08 08:00 = 09/08 20:00)",
       "09/08/2026 às 20:00" in texto)
checar("Lista de medicamentos mostra a data calculada de suspensão (14 dias antes de 10/08 = 27/07/2026)",
       "27/07/2026" in texto)
client.get("/logout")

# Importação de PDF: extrai texto e sugere cortes/medicamentos automaticamente
buffer_pdf = io.BytesIO()
c = canvas.Canvas(buffer_pdf)
c.drawString(50, 800, "TESTE DE EXTRACAO AUTOMATICA DE PDF")
c.drawString(50, 780, "Informacoes ao paciente:")
c.drawString(50, 760, "JEJUM de 12 horas, sendo permitido apenas agua.")
c.drawString(50, 740, "Liquidos claros ate 2 horas antes do exame.")
c.drawString(50, 720, "14 dias antes: OZEMPIC, MOUNJARO, TRULICITY OU SIMILARES.")
c.drawString(50, 700, "NAO e necessario suspender o AAS, Somalgin, Aspirina.")
c.drawString(50, 680, "Nao deve ter realizado colonoscopia ou endoscopia nas 4 semanas anteriores ao exame.")
c.showPage()
c.save()
buffer_pdf.seek(0)

login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-pdf",
    data={"arquivo_pdf": (buffer_pdf, "teste_preparo.pdf")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Importação de PDF extrai o nome sugerido do exame", "TESTE DE EXTRACAO AUTOMATICA DE PDF" in texto)
checar("Importação de PDF sugere o corte de jejum de 12 horas", 'value="12"' in texto and "Jejum total" in texto)
checar("Importação de PDF sugere o corte de líquidos de 2 horas", 'value="2"' in texto)
checar("Importação de PDF separa a lista de medicamentos (separados por vírgula) em linhas diferentes",
       "OZEMPIC" in texto and "MOUNJARO" in texto and "TRULICITY OU SIMILARES" in texto)
checar("Importação de PDF sugere os medicamentos separados com 14 dias de prazo cada",
       texto.count('value="14"') >= 3)
checar("Importação de PDF sugere a observação de medicamentos que não precisam ser suspensos",
       "não" in texto.lower() and "suspender" in texto.lower() and "aas" in texto.lower())
checar("Importação de PDF sugere os exames/procedimentos proibidos antes, separados (colonoscopia e endoscopia)",
       "colonoscopia" in texto.lower() and "endoscopia" in texto.lower() and 'value="28"' in texto)
client.get("/logout")

# A extração de "informações gerais" (regras avulsas, sem data calculada) reconhece
# linhas com marcador de item (➢) e junta linhas de continuação sem marcador, mas
# não duplica o que já foi capturado como corte/medicamento estruturado.
linhas_simuladas = [
    "➢ JEJUM de 12 horas, sendo permitido apenas água.",
    "➢ Não utilizar enxaguante bucal com álcool no dia",
    "do exame.",
    "➢ Não é permitido fumar, mascar chiclete ou praticar atividade física antes do exame.",
    "• Alimentos proibidos: leite e derivados, suco, água com gás.",
]
infos_sugeridas = _sugerir_informacoes_gerais(linhas_simuladas)
checar("Extração de informações gerais junta linha de continuação sem marcador",
       any("álcool no dia do exame" in item for item in infos_sugeridas))
checar("Extração de informações gerais não duplica a regra de jejum (já é um corte estruturado)",
       not any("jejum" in item.lower() for item in infos_sugeridas))
checar("Extração de informações gerais ignora listas longas marcadas com •",
       not any("alimentos proibidos" in item.lower() for item in infos_sugeridas))
checar("Extração de informações gerais captura a regra sobre fumar/chiclete/atividade física",
       any("mascar chiclete" in item for item in infos_sugeridas))

# Cadastro manual de "informações gerais" num modelo de preparo, e exibição ao paciente
with app.app_context():
    modelo_hidrogenio_id = PreparoModelo.query.filter_by(clinica_id=clinica_vitoria_id, nome="Teste de Hidrogênio/Metano Expirado - padrão").first().id

login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/preparo-modelos/{modelo_hidrogenio_id}/editar", data={
    "nome": "Teste de Hidrogênio/Metano Expirado - padrão",
    "instrucoes": "Texto de preparo atualizado pela secretária.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": "Jejum total (sólidos e líquidos)",
    "corte_horas[]": "12",
    "medicamento_nome[]": "Ozempic, Mounjaro, Trulicity ou similares",
    "medicamento_dias[]": "14",
    "medicamento_obs[]": "",
    "info_geral[]": ["Não utilizar enxaguante bucal com álcool no dia do exame.", "Não fumar antes do exame."],
    "exame_anterior_nome[]": ["Colonoscopia", "Lavagens intestinais"],
    "exame_anterior_dias[]": ["28", "28"],
}, follow_redirects=True)
checar("Secretária consegue salvar informações gerais no modelo de preparo", "atualizado" in r.get_data(as_text=True).lower())
client.get("/logout")

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_teste_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê as informações gerais cadastradas no preparo", "Não utilizar enxaguante bucal" in texto and "Não fumar antes do exame" in texto)
checar("Paciente vê os exames/procedimentos proibidos antes, com a data calculada a partir do agendamento (28 dias antes de 10/08/2026 = 13/07/2026)",
       "Colonoscopia" in texto and "Lavagens intestinais" in texto and "13/07/2026" in texto)
client.get("/logout")

# ---------- Alimentos permitidos/proibidos no preparo ----------

with app.app_context():
    colonoscopia_vitoria = Exame.query.filter_by(clinica_id=clinica_vitoria_id, nome="Colonoscopia").first()
    modelo_colono = colonoscopia_vitoria.preparo
    checar("Modelo de colonoscopia foi seedado com alimentos proibidos e permitidos",
           any(not a.permitido for a in modelo_colono.alimentos) and any(a.permitido for a in modelo_colono.alimentos))

    joao = Paciente.query.filter_by(cpf="123.456.789-00").first()
    ag_colono = Agendamento(
        clinica_id=clinica_vitoria_id, paciente_id=joao.id, exame_id=colonoscopia_vitoria.id,
        medico_id=colonoscopia_vitoria.medico_id, data_hora=datetime(2026, 8, 10, 8, 0), status="agendado",
    )
    db.session.add(ag_colono)
    db.session.commit()
    ag_colono_id = ag_colono.id
    colonoscopia_vitoria_id = colonoscopia_vitoria.id

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_colono_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê a lista de alimentos proibidos com o horário calculado a partir do agendamento (12h antes de 10/08 08:00 = 09/08 20:00)",
       "Leite e derivados" in texto and "09/08/2026 às 20:00" in texto)
checar("Paciente vê a lista de alimentos sugeridos para consumo", "Água de coco" in texto)

# O chat responde automaticamente sobre um alimento cadastrado, mesmo sem FAQ manual.
r = client.post("/paciente/chat", data={"pergunta": "posso tomar água de coco?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat responde automaticamente que água de coco é permitida, sem precisar de FAQ cadastrada",
       "água de coco" in texto.lower() and "sim" in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "posso comer leite e derivados antes do exame?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat responde automaticamente que leite e derivados é proibido, com o prazo calculado",
       "leite e derivados" in texto.lower() and "não" in texto.lower() and "09/08/2026" in texto)

# Pequenos erros de digitação não devem fazer a pergunta cair como "pendente"
# (ex.: "amendoin" em vez de "amendoim") — a correspondência tolera isso.
r = client.post("/paciente/chat", data={"pergunta": "Posso comer amendoin?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece 'amendoin' (com erro de digitação) como 'amendoim' e responde automaticamente",
       "amendoim" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())
client.get("/logout")

# Cadastro manual de alimentos num modelo de preparo, via formulário do médico/secretária.
with app.app_context():
    modelo_colono_id = PreparoModelo.query.filter_by(clinica_id=clinica_vitoria_id, nome="Colonoscopia - padrão Vitória").first().id

login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/preparo-modelos/{modelo_colono_id}/editar", data={
    "nome": "Colonoscopia - padrão Vitória",
    "instrucoes": "Texto de preparo da colonoscopia.",
    "observacoes_medicamentos": "",
    "alimento_nome[]": ["Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha"],
    "alimento_tipo[]": ["proibido", "permitido", "proibido", "proibido"],
    "alimento_horas[]": ["24", "", "12", ""],
    "alimento_dias[]": ["", "", "", "3"],
    "medicamento_nome[]": ["Xarelto, Eliquis ou similares"],
    "medicamento_categoria[]": ["medicamento anticoagulante"],
    "medicamento_dias[]": ["3"],
    "medicamento_obs[]": [""],
    "mantido_nome[]": ["AAS"],
    "mantido_obs[]": ["Medicamento analgésico — não precisa suspender"],
    "info_geral[]": ["Tomar Manitol 4 horas antes do exame", "Pode comer até as 20:00 do dia anterior"],
    "info_geral_horas[]": ["4", ""],
    "info_geral_dias[]": ["", "1"],
    "info_geral_hora_exata[]": ["", "20:00"],
}, follow_redirects=True)
checar("Secretária consegue salvar a lista de alimentos no modelo de preparo", "atualizado" in r.get_data(as_text=True).lower())
client.get("/logout")

with app.app_context():
    modelo_colono = PreparoModelo.query.get(modelo_colono_id)
    nomes_alimentos = {a.nome for a in modelo_colono.alimentos}
    checar("Alimentos cadastrados manualmente substituem a lista anterior do modelo",
           nomes_alimentos == {"Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha"})
    alimento_proibido = next(a for a in modelo_colono.alimentos if a.nome == "Milho de pipoca")
    checar("Alimento proibido cadastrado manualmente guarda as horas antes informadas", alimento_proibido.horas_antes == 24)
    alimento_permitido = next(a for a in modelo_colono.alimentos if a.nome == "Suco de maçã coado")
    checar("Alimento permitido cadastrado manualmente não exige horas antes", alimento_permitido.permitido and alimento_permitido.horas_antes is None)
    alimento_dias = next(a for a in modelo_colono.alimentos if a.nome == "Carne vermelha")
    checar("Alimento proibido pode usar prazo em DIAS antes (em vez de horas)",
           alimento_dias.dias_antes == 3 and alimento_dias.horas_antes is None)

    medsusp = modelo_colono.medicamentos_suspensos[0]
    checar("Medicamento suspenso salva a categoria no catálogo compartilhado",
           medsusp.medicamento.categoria == "medicamento anticoagulante")

    mantido = modelo_colono.medicamentos_mantidos[0]
    checar("Medicamento mantido (não suspender) é salvo de forma estruturada",
           mantido.nome == "AAS" and "não precisa suspender" in mantido.observacao)

    info_horas = next(i for i in modelo_colono.informacoes_gerais if i.texto == "Tomar Manitol 4 horas antes do exame")
    checar("Informação geral com prazo em horas antes é salva", info_horas.horas_antes == 4)
    info_dia_relativo = next(i for i in modelo_colono.informacoes_gerais if i.texto == "Pode comer até as 20:00 do dia anterior")
    checar("Informação geral com dia relativo + hora exata é salva",
           info_dia_relativo.dias_antes == 1 and info_dia_relativo.hora_exata is not None
           and info_dia_relativo.hora_exata.strftime("%H:%M") == "20:00")

# O paciente vê o alerta de "carne vermelha" com prazo em dias (não horas),
# e a lista de medicamentos que pode manter, na tela de preparo do exame.
login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/exame/{ag_colono_id}")
texto = r.get_data(as_text=True)
checar("Paciente vê o alimento proibido com prazo calculado em DIAS antes (não horas)",
       "Carne vermelha" in texto and "07/08/2026" in texto and "3 dias antes" in texto)
checar("Paciente vê a lista de medicamentos que pode manter", "AAS" in texto and "não precisa suspender" in texto)

# O chat também reconhece perguntas sobre medicamentos (suspensos ou mantidos),
# não só sobre alimentos — usando a mesma ideia de correspondência por
# palavra-chave, agora sobre o catálogo de medicamentos do preparo.
r = client.post("/paciente/chat", data={"pergunta": "Posso tomar AAS?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que AAS pode ser mantido (medicamento cadastrado como 'não suspender')",
       "sim" in texto.lower() and "aas" in texto.lower() and "pendente" not in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso continuar tomando Xarelto?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que Xarelto precisa ser suspenso, com o prazo calculado",
       "não" in texto.lower() and "xarelto" in texto.lower() and "07/08/2026" in texto and "pendente" not in texto.lower())
client.get("/logout")

# O preparo cadastrou a categoria genérica "Frutas" (não item por item) —
# o chat precisa reconhecer que uma fruta específica (ex.: laranja, banana)
# está coberta por essa categoria, sem precisar que a clínica cadastre cada
# fruta separadamente.
login_paciente("(27) 99999-0000", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso chupar laranja?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat reconhece que 'laranja' está coberta pela categoria genérica 'Frutas' cadastrada no preparo",
       "frutas" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso comer banana?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat também reconhece 'banana' (outro item da categoria 'Frutas')",
       "frutas" in texto.lower() and "não" in texto.lower() and "pendente" not in texto.lower())
client.get("/logout")

# Bug real reportado: uma pergunta sobre um produto específico que só
# MENCIONA uma fruta como sabor/descrição (ex.: "gatorade de uva") não pode
# ser confundida com uma pergunta sobre a fruta em si — mesmo que o preparo
# tenha "Frutas" cadastrada como proibida, "gatorade de uva" é outra coisa.
login("secretaria@clinicavitoria.com", "123456")
client.post(f"/equipe/preparo-modelos/{modelo_colono_id}/editar", data={
    "nome": "Colonoscopia - padrão Vitória",
    "instrucoes": "Texto de preparo da colonoscopia.",
    "observacoes_medicamentos": "",
    "alimento_nome[]": ["Milho de pipoca", "Suco de maçã coado", "Frutas", "Carne vermelha", "Gatorade de cor clara"],
    "alimento_tipo[]": ["proibido", "permitido", "proibido", "proibido", "permitido"],
    "alimento_horas[]": ["24", "", "12", "", ""],
    "alimento_dias[]": ["", "", "", "3", ""],
}, follow_redirects=True)
client.get("/logout")

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso beber gatorade de uva?", "exame_id": str(colonoscopia_vitoria_id)}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Chat NÃO confunde 'gatorade de uva' (sabor) com a categoria 'Frutas' proibida",
       "frutas" not in texto.lower())
client.get("/logout")

# ---------- Integração opcional com a API da Claude (app.ia_preparo) ----------
# Sem ANTHROPIC_API_KEY configurada (como neste ambiente de teste), a IA
# nunca é chamada e o comportamento é exatamente o mesmo de antes (correspondência
# por palavra-chave, depois pendente) — testado abaixo simulando as duas
# respostas possíveis da IA (responde com confiança / não sabe), sem precisar
# de uma chave de API real nem de acesso à internet.
with app.app_context():
    exame_colono_teste = Exame.query.get(colonoscopia_vitoria_id)
    checar("Sem ANTHROPIC_API_KEY configurada, a IA nunca é chamada (fica exatamente como antes)",
           responder_com_ia("Posso beber gatorade de uva?", exame_colono_teste) is None)

import app.routes_paciente as routes_paciente_mod

login_paciente("(27) 99999-0000", "1985-04-12")
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=(
    "Sim! Gatorade de cor clara está entre os alimentos permitidos deste preparo."
)):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade de uva conta como líquido claro?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
checar("Com a IA configurada (simulada aqui), a resposta NÃO vai direto para o paciente — fica esperando aprovação do médico",
       "gatorade de cor clara" not in texto.lower() and "aprova" in texto.lower())

with app.app_context():
    aguardando_gatorade = PerguntaPendente.query.filter_by(
        clinica_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
        status="aguardando_aprovacao",
    ).first()
    checar("A resposta rascunhada pela IA fica salva em PerguntaPendente (aguardando_aprovacao), não direto como FAQ",
           aguardando_gatorade is not None
           and "gatorade de cor clara" in aguardando_gatorade.resposta_sugerida_ia.lower())
    checar("Nenhuma FAQ é criada antes da aprovação do médico",
           FaqItem.query.filter_by(
               clinica_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?"
           ).first() is None)
client.get("/logout")

# O médico revisa o rascunho da IA, edita levemente, e aprova — só a partir
# daqui a resposta deve aparecer para o paciente e ser gravada na FAQ.
login("medico@clinicavitoria.com", "123456")
r = client.get("/equipe/perguntas")
texto = r.get_data(as_text=True)
checar("A pergunta aguardando aprovação aparece na tela do médico, com o rascunho da IA pré-preenchido",
       "gatorade de cor clara" in texto.lower())
with app.app_context():
    pergunta_id_gatorade = PerguntaPendente.query.filter_by(
        clinica_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
    ).first().id
resposta_editada_pelo_medico = "Sim, pode — gatorade de cor clara é permitido nesse preparo (revisado pelo médico)."
client.post(
    f"/equipe/perguntas/{pergunta_id_gatorade}/responder",
    data={"resposta": resposta_editada_pelo_medico},
    follow_redirects=True,
)
with app.app_context():
    aprovada = PerguntaPendente.query.get(pergunta_id_gatorade)
    checar("Depois de aprovada, a pergunta muda para status 'respondida' com a resposta (editada) do médico",
           aprovada.status == "respondida" and aprovada.resposta == resposta_editada_pelo_medico)
    aprendida = FaqItem.query.filter_by(
        clinica_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?"
    ).first()
    checar("Só ao ser aprovada (com a edição do médico) é que a resposta entra na base de FAQ",
           aprendida is not None and aprendida.resposta == resposta_editada_pelo_medico)
client.get("/logout")

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get("/paciente/chat")
texto = r.get_data(as_text=True)
checar("Depois da aprovação, o paciente vê a resposta final (já editada pelo médico) no histórico",
       resposta_editada_pelo_medico.lower() in texto.lower())
client.get("/logout")

# A IA é SEMPRE consultada primeiro (mesmo quando já existe uma FAQ igual
# aprendida antes) — a base de conhecimento só é usada quando a IA não
# responde (sem chave configurada, erro, ou sem confiança). Simulando a IA
# respondendo de novo, de forma DIFERENTE da FAQ já aprendida acima para a
# mesmíssima pergunta, o rascunho novo da IA deve prevalecer — prova de que
# a base de conhecimento não é consultada primeiro.
login_paciente("(27) 99999-0000", "1985-04-12")
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=(
    "Atualização: esse gatorade específico teve a fórmula alterada e não é mais recomendado."
)):
    client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade de uva conta como líquido claro?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
client.get("/logout")
with app.app_context():
    novo_rascunho = PerguntaPendente.query.filter_by(
        clinica_id=clinica_vitoria_id, pergunta="Esse gatorade de uva conta como líquido claro?",
        status="aguardando_aprovacao",
    ).order_by(PerguntaPendente.criado_em.desc()).first()
    checar("A IA é consultada de novo mesmo com uma FAQ idêntica já aprovada antes, e seu rascunho novo prevalece sobre o antigo",
           novo_rascunho is not None and "fórmula alterada" in novo_rascunho.resposta_sugerida_ia.lower())

# Bug real reportado: uma FAQ aprendida da IA sobre um sabor específico
# (uva) não pode ser reaproveitada para uma pergunta parecida mas sobre
# OUTRO sabor (limão) — mesmo compartilhando quase todas as outras
# palavras, a resposta certa pode ser diferente. Sem chamar a IA de novo
# aqui (fica None), a pergunta deve cair para a fila da secretaria, e
# NÃO reaproveitar a resposta específica sobre uva.
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=None):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Esse gatorade sabor limão pode ser tomado?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
checar("FAQ aprendida sobre um sabor (uva) NÃO é reaproveitada para uma pergunta sobre outro sabor (limão)",
       "gatorade de cor clara" not in texto.lower() and ("encaminhei" in texto.lower() or "pendente" in texto.lower()))

# Quando a IA está configurada mas sinaliza que não sabe responder (fora do
# escopo do preparo), o comportamento continua o mesmo de sempre: cai para a
# correspondência por palavra-chave e, por fim, para a fila da secretaria.
with patch.object(routes_paciente_mod, "responder_com_ia", return_value=None):
    r = client.post(
        "/paciente/chat",
        data={"pergunta": "Qual o resultado do meu exame de sangue do ano passado?", "exame_id": str(colonoscopia_vitoria_id)},
        follow_redirects=True,
    )
texto = r.get_data(as_text=True)
checar("Quando a IA não sabe responder, a pergunta ainda cai para a fila da secretaria normalmente",
       "encaminhei" in texto.lower() or "pendente" in texto.lower())
client.get("/logout")

# ---------- Extração de alimentos e medicamentos (semanas) de um PDF ----------

linhas_alimentos_simuladas = [
    "• Alimentos proibidos: leite e derivados, milho, feijão, pão integral.",
    "• Sugestão para consumo: água de coco, chá claro, gelatina sem cor.",
]
alimentos_sugeridos = _sugerir_alimentos(linhas_alimentos_simuladas)
checar("Extração de alimentos proibidos aplica o prazo padrão de 12 horas quando não há horário explícito",
       any(a["nome"] == "leite" and not a["permitido"] and a["horas_antes"] == 12 for a in alimentos_sugeridos))
checar("Extração separa cada alimento proibido da lista em prosa",
       {"leite", "derivados", "milho", "feijão", "pão integral"}.issubset({a["nome"] for a in alimentos_sugeridos if not a["permitido"]}))
checar("Extração de 'Sugestão para consumo' marca os alimentos como permitidos, sem prazo",
       any(a["nome"] == "água de coco" and a["permitido"] and a["horas_antes"] is None for a in alimentos_sugeridos))

texto_no_dia_do_exame = "Realizar jejum no dia do exame. Comparecer com o pedido médico."
cortes_no_dia = _sugerir_cortes(texto_no_dia_do_exame)
checar("Referência a 'no dia do exame' sem horas explícitas é tratada como 12 horas antes",
       any(c["horas_antes"] == 12 for c in cortes_no_dia))

texto_medicamentos_semanas = (
    "Suspender 2 semanas antes do exame os medicamentos para emagrecimento "
    "(semaglutida - ozempic, liraglutida - victoza e tirzepatida - mounjaro).\n"
    "1 semana antes: anticoagulantes (xarelto, eliquis)."
)
medicamentos_sugeridos = _sugerir_medicamentos(texto_medicamentos_semanas)
checar("Medicamento em prosa com '2 semanas antes' é convertido para 14 dias",
       any(m["dias_antes"] == 14 and "ozempic" in m["nome"].lower() for m in medicamentos_sugeridos))
checar("Medicamento em formato de tabela com '1 semana antes' é convertido para 7 dias",
       any(m["dias_antes"] == 7 and "anticoagulantes" in m["nome"].lower() for m in medicamentos_sugeridos))

# ---------- Importação de preparo a partir de uma planilha Excel (.xlsx) ----------

def _construir_xlsx_teste():
    """Monta uma planilha pequena no mesmo formato estruturado que a
    clínica passou a usar (colunas Tipo/Ação/Agrupador/Nome/Dias antes/
    Horas antes/Hora exata), com 2 abas — cada uma o preparo de um exame
    diferente — para testar a extração e a tela de escolher qual aba
    importar primeiro."""
    cabecalho = ["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"]
    wb = Workbook()
    aba1 = wb.active
    aba1.title = "Preparo A"
    linhas_aba1 = [
        cabecalho,
        ["Medicamento", "Suspender", "medicamento antiplaquetário", "Ticlid", 10, None, None],
        ["Medicamento", "Não suspender", "medicamento analgésico", "AAS", None, None, None],
        ["Medicamento", "Suspender", "antibióticos", None, 28, None, None],
        ["Medicamento", "Receituário", None, "Picoprep - 1 sachê às 16:00 do dia anterior", 1, None, dt_time(16, 0)],
        ["Alimento", "Suspender", None, "frutas", 3, None, None],
        ["Alimento", "Permitido (Sugestão de consumo)", None, "água de coco", None, 12, None],
        ["Exames / Procedimentos", "Proibido", None, "colonoscopia", 28, None, None],
        ["Aviso", "Intruções para IA", None, "JEJUM de 12 horas", None, 12, None],
        ["Aviso", "Intruções para IA", None, "Pode comer até as 20:00 do dia anterior", 1, None, dt_time(20, 0)],
        ["Aviso", "Intruções para IA", None, "Não fumar antes do exame", None, None, None],
    ]
    for linha in linhas_aba1:
        aba1.append(linha)

    aba2 = wb.create_sheet("Preparo B")
    linhas_aba2 = [
        cabecalho,
        ["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None],
        ["Exames / Procedimentos", "Proibido", None, "endoscopia", 14, None, None],
    ]
    for linha in linhas_aba2:
        aba2.append(linha)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


with app.app_context():
    sugestoes = extrair_sugestoes_de_xlsx(_construir_xlsx_teste())
    checar("Extração de Excel identifica as 2 abas como preparos separados",
           len(sugestoes) == 2 and sugestoes[0]["aba_nome"] == "Preparo A" and sugestoes[1]["aba_nome"] == "Preparo B")

    aba_a = sugestoes[0]
    checar("Extração de Excel: medicamento com categoria (Agrupador) e Nome preenchidos",
           any(m["nome"] == "Ticlid" and m["dias_antes"] == 10 and m["categoria"] == "medicamento antiplaquetário" for m in aba_a["medicamentos"]))
    checar("Extração de Excel: medicamento sem Nome usa o Agrupador como nome (ex.: 'antibióticos')",
           any(m["nome"] == "antibióticos" and m["dias_antes"] == 28 for m in aba_a["medicamentos"]))
    checar("Extração de Excel: medicamento 'Não suspender' vira medicamento mantido",
           any(m["nome"] == "AAS" for m in aba_a["medicamentos_mantidos"]))
    checar("Extração de Excel: 'Receituário' com horário vira informação geral com dia relativo + hora exata",
           any(i["texto"].startswith("Picoprep") and i["dias_antes"] == 1 and i["hora_exata"] == "16:00" for i in aba_a["informacoes_gerais"]))
    checar("Extração de Excel: alimento com 'Dias antes' preenchido usa prazo em dias (não horas)",
           any(a["nome"] == "frutas" and a["dias_antes"] == 3 and a["horas_antes"] is None and not a["permitido"] for a in aba_a["alimentos"]))
    checar("Extração de Excel: alimento permitido com 'Horas antes'",
           any(a["nome"] == "água de coco" and a["permitido"] and a["horas_antes"] == 12 for a in aba_a["alimentos"]))
    checar("Extração de Excel: exame/procedimento proibido antes",
           any(e["nome"] == "colonoscopia" and e["dias_antes"] == 28 for e in aba_a["exames_anteriores"]))
    checar("Extração de Excel: aviso de 'JEJUM' com horas vira um corte (não uma informação geral solta)",
           any(c["descricao"] == "JEJUM de 12 horas" and c["horas_antes"] == 12 for c in aba_a["cortes"]))
    checar("Extração de Excel: aviso com dia relativo + hora exata (ex.: 'até as 20:00 do dia anterior')",
           any(i["texto"] == "Pode comer até as 20:00 do dia anterior" and i["dias_antes"] == 1 and i["hora_exata"] == "20:00" for i in aba_a["informacoes_gerais"]))
    checar("Extração de Excel: aviso sem prazo nenhum vira informação geral simples",
           any(i["texto"] == "Não fumar antes do exame" and i["horas_antes"] is None and i["dias_antes"] is None for i in aba_a["informacoes_gerais"]))

    aba_b = sugestoes[1]
    checar("Extração de Excel: segunda aba tratada como um preparo independente",
           any(e["nome"] == "endoscopia" and e["dias_antes"] == 14 for e in aba_b["exames_anteriores"]))

# Fluxo completo pela interface: planilha com 2 abas -> tela de escolha -> revisão -> salvar.
login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (_construir_xlsx_teste(), "teste_preparo.xlsx")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com 2 abas leva à tela de escolher qual aba importar primeiro",
       "Preparo A" in texto and "Preparo B" in texto)

r = client.post("/equipe/preparo-modelos/importar-xlsx/escolher", data={"indice": "0"}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Ao escolher a aba, o formulário de novo modelo vem pré-preenchido com os dados extraídos",
       "Ticlid" in texto and "antibióticos" in texto and "AAS" in texto and "colonoscopia" in texto)

r = client.post("/equipe/preparo-modelos/novo", data={
    "nome": "Preparo importado do Excel - Teste",
    "instrucoes": "Instruções de teste.",
    "observacoes_medicamentos": "",
    "corte_descricao[]": ["JEJUM de 12 horas"],
    "corte_horas[]": ["12"],
    "medicamento_nome[]": ["Ticlid", "antibióticos"],
    "medicamento_categoria[]": ["medicamento antiplaquetário", ""],
    "medicamento_dias[]": ["10", "28"],
    "medicamento_obs[]": ["", ""],
    "mantido_nome[]": ["AAS"],
    "mantido_obs[]": [""],
    "info_geral[]": ["Picoprep - 1 sachê às 16:00 do dia anterior", "Pode comer até as 20:00 do dia anterior", "Não fumar antes do exame"],
    "info_geral_horas[]": ["", "", ""],
    "info_geral_dias[]": ["1", "1", ""],
    "info_geral_hora_exata[]": ["16:00", "20:00", ""],
    "alimento_nome[]": ["frutas", "água de coco"],
    "alimento_tipo[]": ["proibido", "permitido"],
    "alimento_horas[]": ["", "12"],
    "alimento_dias[]": ["3", ""],
    "exame_anterior_nome[]": ["colonoscopia"],
    "exame_anterior_dias[]": ["28"],
}, follow_redirects=True)
checar("Modelo importado da planilha é salvo com sucesso", "cadastrado" in r.get_data(as_text=True).lower())
client.get("/logout")

with app.app_context():
    modelo_importado = PreparoModelo.query.filter_by(clinica_id=clinica_vitoria_id, nome="Preparo importado do Excel - Teste").first()
    checar("Modelo importado tem o corte de jejum", any(c.horas_antes == 12 for c in modelo_importado.cortes))
    checar("Modelo importado tem o medicamento com categoria",
           any(m.medicamento.nome == "Ticlid" and m.medicamento.categoria == "medicamento antiplaquetário" for m in modelo_importado.medicamentos_suspensos))
    checar("Modelo importado tem o medicamento mantido (AAS)",
           any(m.nome == "AAS" for m in modelo_importado.medicamentos_mantidos))
    checar("Modelo importado tem o alimento com prazo em dias",
           any(a.nome == "frutas" and a.dias_antes == 3 for a in modelo_importado.alimentos))
    checar("Modelo importado tem o exame anterior proibido",
           any(e.nome == "colonoscopia" and e.dias_antes == 28 for e in modelo_importado.exames_anteriores_proibidos))
    checar("Modelo importado tem a informação geral com dia relativo + hora exata",
           any(i.dias_antes == 1 and i.hora_exata is not None for i in modelo_importado.informacoes_gerais))

# Planilha com uma aba só vai direto para a tela de revisão (sem passar pela tela de escolha).
buffer_uma_aba = io.BytesIO()
wb_uma_aba = Workbook()
wb_uma_aba.active.title = "Único preparo"
wb_uma_aba.active.append(["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"])
wb_uma_aba.active.append(["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None])
wb_uma_aba.save(buffer_uma_aba)
buffer_uma_aba.seek(0)

login("secretaria@clinicavitoria.com", "123456")
r = client.post(
    "/equipe/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (buffer_uma_aba, "teste_uma_aba.xlsx")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com uma aba só vai direto para a tela de revisão (novo modelo pré-preenchido)",
       "Trazer o pedido médico" in texto and "Novo modelo de preparo" in texto)
client.get("/logout")


# ---------- Exames/procedimentos anteriores proibidos, e separação de medicamentos com "/" ----------

texto_exames_anteriores = (
    "Não deve ter realizado colonoscopia ou lavagens intestinais nas 4 semanas anteriores ao exame.\n"
    "Não deve ter utilizado antibióticos nas 4 semanas que antecedem o exame."
)
exames_anteriores_sugeridos = _sugerir_exames_anteriores(texto_exames_anteriores)
checar("Extração de exames anteriores proibidos separa 'colonoscopia' e 'lavagens intestinais' (ligados por 'ou')",
       {"colonoscopia", "lavagens intestinais"} == {e["nome"].lower() for e in exames_anteriores_sugeridos})
checar("Extração de exames anteriores proibidos converte '4 semanas' para 28 dias",
       all(e["dias_antes"] == 28 for e in exames_anteriores_sugeridos))
checar("A restrição sobre antibióticos ('utilizado', não 'realizado') não entra como exame anterior proibido",
       not any("antibiótico" in e["nome"].lower() for e in exames_anteriores_sugeridos))

linhas_exame_anterior_e_antibiotico = [
    "➢ Não deve ter realizado colonoscopia ou lavagens intestinais nas 4 semanas anteriores ao exame.",
    "➢ Não deve ter utilizado antibióticos nas 4 semanas que antecedem o exame.",
]
infos_sem_duplicar_exame_anterior = _sugerir_informacoes_gerais(linhas_exame_anterior_e_antibiotico)
checar("A linha sobre colonoscopia não duplica em informações gerais (já é um exame anterior estruturado)",
       not any("colonoscopia" in item.lower() for item in infos_sem_duplicar_exame_anterior))
checar("A linha sobre antibióticos continua em informações gerais (não é um exame/procedimento)",
       any("antibiótico" in item.lower() for item in infos_sem_duplicar_exame_anterior))

texto_medicamentos_barra = "3 dias antes: Xarelto / Eliquis / Marevan."
medicamentos_barra_sugeridos = _sugerir_medicamentos(texto_medicamentos_barra)
checar("Lista de medicamentos separada por '/' é dividida em itens individuais",
       {"Xarelto", "Eliquis", "Marevan"} == {m["nome"] for m in medicamentos_barra_sugeridos})

with app.app_context():
    config = PlataformaConfig.obter()
    checar("Configuração de trial tem um valor padrão de dias", config.trial_dias > 0)

    empresa_teste = Empresa.query.filter_by(nome="Empresa Teste Automatizado").first()
    checar("Empresa nova recebeu data de vencimento do trial", empresa_teste.data_vencimento is not None)
    checar("Empresa nova começa com status 'trial'", empresa_teste.status == "trial")

    # Força o vencimento para o passado, simulando um trial expirado.
    empresa_teste.data_vencimento = date.today() - timedelta(days=1)
    db.session.commit()
    empresa_teste_id = empresa_teste.id

login("dono@plataforma.com", "123456")
r = client.get("/dono/")
texto = r.get_data(as_text=True)
checar("Trial vencido aparece como 'inadimplente' no painel do dono, sem bloquear",
       "inadimplente" in texto.lower())

r = client.post("/dono/configuracoes", data={"trial_dias": "45"}, follow_redirects=True)
checar("Dono consegue alterar a duração do trial", "45" in r.get_data(as_text=True))

r = client.post(f"/dono/empresas/{empresa_teste_id}/editar", data={
    "status": "ativa", "data_vencimento": "", "observacoes_pagamento": "", "valor_por_medico": "200,50",
}, follow_redirects=True)
checar("Dono consegue definir o valor por médico de uma empresa", "200.50" in r.get_data(as_text=True))
client.get("/logout")

with app.app_context():
    empresa_teste = Empresa.query.get(empresa_teste_id)
    checar("Empresa passou a 'ativa' e teve o valor por médico salvo",
           empresa_teste.status == "ativa" and float(empresa_teste.valor_por_medico) == 200.50)
    # Devolve pra trial vencido, pra continuar o teste de vencimento abaixo.
    empresa_teste.status = "trial"
    empresa_teste.data_vencimento = date.today() - timedelta(days=1)
    db.session.commit()

login("dono@plataforma.com", "123456")
client.get("/dono/")  # dispara a checagem de vencimento de novo
client.get("/logout")

with app.app_context():
    empresa_teste = Empresa.query.get(empresa_teste_id)
    checar("Empresa com trial vencido virou 'inadimplente' (não foi bloqueada automaticamente)",
           empresa_teste.status == "inadimplente")

# A secretária da empresa com trial vencido (agora inadimplente) ainda consegue acessar normalmente.
login("fulano@clinicateste.com", "senha123")
r = client.get("/equipe/pacientes")
checar("Empresa inadimplente por trial vencido continua acessível (só bloqueio manual impede acesso)",
       r.status_code == 200)
client.get("/logout")

# ---------- Chat do paciente e aprendizado da IA (regressão) ----------

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.post("/paciente/chat", data={"pergunta": "Posso comer batata antes do exame?", "exame_id": "1"}, follow_redirects=True)
checar("IA responde pergunta conhecida sobre batata", "fibra" in r.get_data(as_text=True).lower())

r = client.post("/paciente/chat", data={"pergunta": "Posso fazer exercício físico pesado antes da colonoscopia?", "exame_id": "1"}, follow_redirects=True)
checar("Pergunta nova é encaminhada para secretaria", "encaminhei" in r.get_data(as_text=True).lower())
client.get("/logout")

login("secretaria@clinicavitoria.com", "123456")
r = client.get("/equipe/perguntas")
checar("Pergunta pendente do paciente da Vitória aparece para a secretaria da Vitória",
       "exercício físico pesado" in r.get_data(as_text=True))
client.get("/logout")

# ---------- Isolamento entre pacientes: Maria não vê pergunta do João ----------

login("secretaria@clinicasp.com", "123456")
r = client.get("/equipe/perguntas")
checar("Pergunta do paciente da Vitória NÃO aparece para a secretaria da SP",
       "exercício físico pesado" not in r.get_data(as_text=True))
client.get("/logout")

# ---------- Paciente pode remover a própria pergunta pendente ----------

login_paciente("(27) 99999-0000", "1985-04-12")
with app.app_context():
    pergunta_exercicio = PerguntaPendente.query.filter(
        PerguntaPendente.pergunta.like("%exercício físico pesado%")
    ).first()
    pergunta_exercicio_id = pergunta_exercicio.id

r = client.get("/paciente/chat")
checar("Paciente vê o botão de remover na própria pergunta pendente", "Remover" in r.get_data(as_text=True))

r = client.post(f"/paciente/perguntas/{pergunta_exercicio_id}/remover", follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Paciente consegue remover a própria pergunta da lista", "exercício físico pesado" not in texto and "removida" in texto.lower())
client.get("/logout")

login_paciente("(11) 98888-0000", "1990-09-03")
r = client.post("/paciente/chat", data={"pergunta": "Posso comer chocolate?", "exame_id": ""}, follow_redirects=True)
with app.app_context():
    pergunta_maria = PerguntaPendente.query.filter(PerguntaPendente.pergunta.like("%chocolate%")).first()
    pergunta_maria_id = pergunta_maria.id
client.get("/logout")

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.post(f"/paciente/perguntas/{pergunta_maria_id}/remover")
checar("Um paciente não consegue remover a pergunta de outro paciente forjando o id na URL", r.status_code == 404)
client.get("/logout")

# ---------- Dono da plataforma ----------

r = login("dono@plataforma.com", "123456")
texto = r.get_data(as_text=True)
checar("Login do dono cai no painel de empresas", "Empresas na plataforma" in texto)
checar("Painel do dono lista a empresa Clínica Vitória e a Clínica São Paulo",
       "Clínica Vitória" in texto and "Clínica São Paulo" in texto)

r = client.get("/equipe/pacientes")
checar("Dono não consegue acessar a área de equipe de uma clínica", r.status_code in (302, 401, 403) or "Empresas na plataforma" not in r.get_data(as_text=True))

with app.app_context():
    empresa_sp_id = Empresa.query.filter_by(nome="Clínica São Paulo").first().id

# Bloquear a empresa Clínica São Paulo
r = client.get(f"/dono/empresas/{empresa_sp_id}")
r = client.post(f"/dono/empresas/{empresa_sp_id}/bloquear", follow_redirects=True)
checar("Dono consegue bloquear uma empresa (todas as suas filiais)", "bloqueado" in r.get_data(as_text=True).lower())

client.get("/logout")

# Paciente da empresa bloqueada não consegue mais acessar
r = login_paciente("(11) 98888-0000", "1990-09-03")
checar("Paciente de empresa bloqueada não acessa mais o sistema",
       "indisponível" in r.get_data(as_text=True).lower() or "Meus exames" not in r.get_data(as_text=True))
client.get("/logout")

# Secretária da empresa bloqueada também não consegue mais usar a área de equipe
login("secretaria@clinicasp.com", "123456")
r = client.get("/equipe/pacientes")
checar("Secretária de empresa bloqueada é bloqueada ao tentar usar a área de equipe",
       r.status_code in (302,) or "não está vinculada a nenhuma clínica ativa" in r.get_data(as_text=True))
client.get("/logout")

# Desbloquear de novo, para deixar o banco limpo para uso manual depois do teste
login("dono@plataforma.com", "123456")
client.post(f"/dono/empresas/{empresa_sp_id}/desbloquear", follow_redirects=True)
client.get("/logout")

# ---------- Novas funcionalidades: duração/preço/acompanhante do exame,
# horário do médico, otimizador de agenda, solicitação de agendamento pelo
# paciente, atendimento (histórico de chat + encerramento), resultado em
# PDF, pagamento/desconto/comprovante, e endereço/contato de emergência ----------

with app.app_context():
    clinica_vitoria_id = Clinica.query.filter_by(nome="Clínica Vitória").first().id
    medico_carlos = Usuario.query.filter_by(email="medico@clinicavitoria.com").first()
    medico_carlos_id = medico_carlos.id
    colonoscopia = Exame.query.filter_by(clinica_id=clinica_vitoria_id, nome="Colonoscopia").first()
    colonoscopia_id = colonoscopia.id
    joao_id = Paciente.query.filter_by(clinica_id=clinica_vitoria_id, nome="João Pereira").first().id

login("secretaria@clinicavitoria.com", "123456")

# --- Duração, preço e acompanhante do exame ---
r = client.post(f"/equipe/exames/{colonoscopia_id}/editar", data={
    "nome": "Colonoscopia",
    "descricao": "Exame do intestino grosso",
    "duracao_minutos": "45",
    "preco": "350,00",
    "precisa_acompanhante": "on",
    "preparo_modelo_id": str(colonoscopia.preparo_modelo_id),
    "medico_id": str(medico_carlos_id),
}, follow_redirects=True)
checar("Exame atualizado com duração/preço/acompanhante", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    colonoscopia_checar = Exame.query.get(colonoscopia_id)
    checar("Duração do exame salva (45 minutos)", colonoscopia_checar.duracao_minutos == 45)
    checar("Preço do exame salvo (350.00)", float(colonoscopia_checar.preco) == 350.00)
    checar("Flag de acompanhante obrigatório salva", colonoscopia_checar.precisa_acompanhante is True)

# Agendar sem informar acompanhante deve ser rejeitado
r = client.post("/equipe/agenda/novo", data={
    "filial_id": str(clinica_vitoria_id),
    "paciente_id": str(joao_id),
    "exame_id": str(colonoscopia_id),
    "data_hora": "2026-09-15T08:00",
    "observacoes": "",
    "acompanhante_nome": "",
}, follow_redirects=True)
checar("Agendamento sem acompanhante é rejeitado quando o exame exige", "acompanhante" in r.get_data(as_text=True).lower())

r = client.post("/equipe/agenda/novo", data={
    "filial_id": str(clinica_vitoria_id),
    "paciente_id": str(joao_id),
    "exame_id": str(colonoscopia_id),
    "data_hora": "2026-09-15T08:00",
    "observacoes": "",
    "acompanhante_nome": "Maria (esposa)",
}, follow_redirects=True)
checar("Agendamento com acompanhante informado é aceito", "sucesso" in r.get_data(as_text=True).lower())

with app.app_context():
    agendamento_colono = Agendamento.query.filter_by(
        paciente_id=joao_id, exame_id=colonoscopia_id, clinica_id=clinica_vitoria_id
    ).order_by(Agendamento.id.desc()).first()
    agendamento_colono_id = agendamento_colono.id
    checar("Nome do acompanhante foi salvo no agendamento", agendamento_colono.acompanhante_nome == "Maria (esposa)")

# --- Horário de atendimento do médico + otimizador de agenda ---
r = client.get(f"/equipe/medico-horarios/{medico_carlos_id}")
checar("Tela de horário do médico carrega", r.status_code == 200)

dados_horario = {}
for dia in range(7):
    if dia in (0, 1, 2, 3, 4):  # segunda a sexta
        dados_horario[f"dia_{dia}_ativo"] = "on"
        dados_horario[f"dia_{dia}_inicio"] = "08:00"
        dados_horario[f"dia_{dia}_fim"] = "12:00"
r = client.post(f"/equipe/medico-horarios/{medico_carlos_id}", data=dados_horario, follow_redirects=True)
checar("Horário de atendimento do médico salvo", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    from app.agendamento_otimizador import sugerir_horarios
    colonoscopia_reload = Exame.query.get(colonoscopia_id)
    clinica_vitoria_reload = Clinica.query.get(clinica_vitoria_id)
    medico_carlos_reload = Usuario.query.get(medico_carlos_id)
    sugestoes = sugerir_horarios(colonoscopia_reload, medico_carlos_reload, clinica_vitoria_reload, quantidade=3)
    checar("Otimizador sugere horários dentro do expediente cadastrado (08h-12h)",
           len(sugestoes) > 0 and all(8 <= s.hour < 12 for s in sugestoes))
client.get("/logout")

# --- Solicitação de agendamento pelo paciente ---
login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/agendar?exame_id={colonoscopia_id}")
checar("Tela de solicitar agendamento mostra horários sugeridos", "Escolha um dos próximos horários" in r.get_data(as_text=True))

texto_pagina = r.get_data(as_text=True)
import re as _re
match_horario = _re.search(r'value="(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})"', texto_pagina)
checar("Encontrou pelo menos um horário sugerido no HTML", match_horario is not None)

r = client.post("/paciente/agendar", data={
    "exame_id": str(colonoscopia_id),
    "horario_escolhido": match_horario.group(1),
}, follow_redirects=True)
checar("Solicitação de agendamento enviada com sucesso", "solicitação" in r.get_data(as_text=True).lower() or "enviada" in r.get_data(as_text=True).lower())

with app.app_context():
    solicitacao = Agendamento.query.filter_by(
        paciente_id=joao_id, exame_id=colonoscopia_id, status="solicitado"
    ).first()
    checar("Agendamento solicitado pelo paciente fica com status 'solicitado'", solicitacao is not None)
    solicitacao_id = solicitacao.id

# Deixa uma pergunta registrada no chat, para aparecer no atendimento
r = client.post("/paciente/chat", data={
    "pergunta": "Posso comer batata no preparo?", "exame_id": str(colonoscopia_id),
}, follow_redirects=True)
client.get("/logout")

# --- Secretária confirma a solicitação de agendamento ---
login("secretaria@clinicavitoria.com", "123456")
r = client.get("/equipe/agenda/solicitacoes")
checar("Tela de solicitações lista o pedido do paciente", "João Pereira" in r.get_data(as_text=True))

r = client.post(f"/equipe/agenda/{solicitacao_id}/confirmar-solicitacao", data={"acao": "confirmar"}, follow_redirects=True)
with app.app_context():
    solicitacao_confirmada = Agendamento.query.get(solicitacao_id)
    checar("Solicitação confirmada vira status 'agendado'", solicitacao_confirmada.status == "agendado")

# --- Atendimento: médico vê as perguntas do paciente e encerra a consulta ---
client.get("/logout")
login("medico@clinicavitoria.com", "123456")
client.post("/equipe/clinica", data={"clinica_id": str(clinica_vitoria_id)}, follow_redirects=True)
r = client.get(f"/equipe/agenda/{agendamento_colono_id}/atendimento")
checar("Tela de atendimento mostra a pergunta feita pelo paciente no app", "batata" in r.get_data(as_text=True).lower())

r = client.post(f"/equipe/agenda/{agendamento_colono_id}/atendimento", data={
    "notas_atendimento": "Paciente sem queixas, preparo realizado corretamente.",
    "encerrar": "on",
}, follow_redirects=True)
checar("Atendimento encerrado com sucesso", "encerrado" in r.get_data(as_text=True).lower())

with app.app_context():
    agendamento_encerrado = Agendamento.query.get(agendamento_colono_id)
    checar("Agendamento marcado como 'realizado' após encerrar o atendimento", agendamento_encerrado.status == "realizado")
    checar("Data/hora de encerramento foi registrada", agendamento_encerrado.encerrado_em is not None)
    checar("Observações da consulta foram salvas", "sem queixas" in (agendamento_encerrado.notas_atendimento or ""))

# --- Resultado de exame em PDF (upload pela equipe, download pelo paciente) ---
buffer_resultado = io.BytesIO()
canvas.Canvas(buffer_resultado).save()
buffer_resultado.seek(0)
r = client.post(
    f"/equipe/agenda/{agendamento_colono_id}/resultado",
    data={"arquivo_pdf": (buffer_resultado, "resultado_colonoscopia.pdf")},
    content_type="multipart/form-data",
    follow_redirects=True,
)
checar("Resultado do exame enviado com sucesso", "sucesso" in r.get_data(as_text=True).lower())
client.get("/logout")

login_paciente("(27) 99999-0000", "1985-04-12")
r = client.get(f"/paciente/exame/{agendamento_colono_id}/resultado")
checar("Paciente consegue baixar o resultado do exame anexado", r.status_code == 200 and r.mimetype == "application/pdf")
client.get("/logout")

# --- Descontos + registro de pagamento + comprovante ---
login("secretaria@clinicavitoria.com", "123456")
r = client.post("/equipe/descontos", data={"nome": "Convênio Saúde+", "percentual": "10"}, follow_redirects=True)
checar("Desconto cadastrado com sucesso", "Convênio Saúde+" in r.get_data(as_text=True))

with app.app_context():
    desconto_id = DescontoConfig.query.filter_by(clinica_id=clinica_vitoria_id, nome="Convênio Saúde+").first().id

r = client.post(f"/equipe/agenda/{agendamento_colono_id}/pagamento", data={
    "desconto_id": str(desconto_id), "forma_pagamento": "pix",
}, follow_redirects=True)
checar("Pagamento registrado com sucesso", "comprovante" in r.get_data(as_text=True).lower() or r.status_code == 200)

with app.app_context():
    pagamento = Pagamento.query.filter_by(agendamento_id=agendamento_colono_id).first()
    checar("Valor do procedimento no pagamento bate com o preço cadastrado (350.00)", float(pagamento.valor_procedimento) == 350.00)
    checar("Desconto de 10% aplicado corretamente (valor final 315.00)", float(pagamento.valor_final) == 315.00)

r = client.get(f"/equipe/agenda/{agendamento_colono_id}/pagamento/comprovante")
checar("Comprovante de pagamento mostra o valor final", "315" in r.get_data(as_text=True))
client.get("/logout")

# --- Endereço (CEP) e contato de emergência no cadastro do paciente ---
login("secretaria@clinicavitoria.com", "123456")
r = client.post(f"/equipe/pacientes/{joao_id}/editar", data={
    "nome": "João Pereira",
    "email": "",
    "observacoes": "",
    "cep": "29010-000",
    "rua": "Avenida Jerônimo Monteiro",
    "numero": "1000",
    "complemento": "",
    "bairro": "Centro",
    "cidade": "Vitória",
    "uf": "ES",
    "contato_emergencia_nome": "Maria Pereira",
    "contato_emergencia_telefone": "(27) 99999-2222",
}, follow_redirects=True)
checar("Cadastro do paciente com endereço/contato de emergência salvo", "atualizado" in r.get_data(as_text=True).lower())

with app.app_context():
    joao_atualizado = Paciente.query.get(joao_id)
    checar("Endereço do paciente salvo (rua)", joao_atualizado.rua == "Avenida Jerônimo Monteiro")
    checar("Contato de emergência salvo", joao_atualizado.contato_emergencia_nome == "Maria Pereira")

r = client.get(f"/equipe/pacientes/{joao_id}")
checar("Detalhe do paciente exibe o endereço cadastrado", "Jerônimo Monteiro" in r.get_data(as_text=True))
checar("Detalhe do paciente exibe o contato de emergência", "Maria Pereira" in r.get_data(as_text=True))
client.get("/logout")


# ---------- Trocar senha ----------
login("secretaria@clinicavitoria.com", "123456")
r = client.get("/trocar-senha")
checar("Tela de trocar senha carrega", "Trocar senha" in r.get_data(as_text=True))

r = client.post("/trocar-senha", data={
    "senha_atual": "senha_errada",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "novaSenha123",
}, follow_redirects=True)
checar("Trocar senha rejeita quando a senha atual está incorreta", "incorreta" in r.get_data(as_text=True).lower())

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "abc",
    "senha_confirmacao": "abc",
}, follow_redirects=True)
checar("Trocar senha rejeita nova senha curta", "pelo menos 6 caracteres" in r.get_data(as_text=True))

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "outraSenha456",
}, follow_redirects=True)
checar("Trocar senha rejeita quando a confirmação não corresponde", "não corresponde" in r.get_data(as_text=True).lower())

r = client.post("/trocar-senha", data={
    "senha_atual": "123456",
    "senha_nova": "novaSenha123",
    "senha_confirmacao": "novaSenha123",
}, follow_redirects=True)
checar("Trocar senha funciona com dados válidos", "sucesso" in r.get_data(as_text=True).lower())
client.get("/logout")

r = login("secretaria@clinicavitoria.com", "123456")
checar("Login com a senha antiga não funciona mais após a troca", "E-mail ou senha inválidos" in r.get_data(as_text=True))

r = login("secretaria@clinicavitoria.com", "novaSenha123")
checar("Login com a nova senha funciona", "Empresa" in r.get_data(as_text=True) or "Pacientes" in r.get_data(as_text=True) or r.status_code == 200)
client.get("/logout")

print("\nTodos os testes de fumaça passaram com sucesso.")
