"""Utilitários compartilhados pelos relatórios (ver app/routes_relatorios.py):
parsing do filtro de período e exportação da mesma tabela em três formatos
(CSV, Excel e PDF), para não duplicar essa lógica em cada relatório."""
import csv
import io
from datetime import date, datetime, timedelta

from flask import Response, request


def periodo_do_filtro(padrao_dias=30):
    """Lê `data_inicio`/`data_fim` da query string (formato AAAA-MM-DD).
    Sem filtro informado, usa os últimos `padrao_dias` dias terminando
    hoje — um período recente e útil por padrão, sem obrigar a pessoa a
    escolher datas toda vez que abre o relatório."""
    hoje = date.today()

    bruto_inicio = request.args.get("data_inicio", "").strip()
    bruto_fim = request.args.get("data_fim", "").strip()

    try:
        data_inicio = datetime.strptime(bruto_inicio, "%Y-%m-%d").date() if bruto_inicio else hoje - timedelta(days=padrao_dias)
    except ValueError:
        data_inicio = hoje - timedelta(days=padrao_dias)

    try:
        data_fim = datetime.strptime(bruto_fim, "%Y-%m-%d").date() if bruto_fim else hoje
    except ValueError:
        data_fim = hoje

    # Evita um período "invertido" (fim antes do início), que geraria uma
    # consulta sempre vazia sem nenhum aviso do motivo.
    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio

    return data_inicio, data_fim


def intervalo_datetime(data_inicio, data_fim):
    """Converte um par de `date` num par de `datetime` cobrindo o dia
    inteiro de `data_fim` (00:00:00 do início até 23:59:59 do fim) — os
    campos de data/hora no banco (`pago_em`, `data_hora`, `criado_em`) são
    DateTime, então filtrar só por `date` deixaria de fora os registros do
    próprio dia final."""
    inicio_dt = datetime.combine(data_inicio, datetime.min.time())
    fim_dt = datetime.combine(data_fim, datetime.max.time())
    return inicio_dt, fim_dt


def _nome_arquivo(base, data_inicio, data_fim, extensao):
    return f"{base}_{data_inicio.isoformat()}_a_{data_fim.isoformat()}.{extensao}"


def exportar_csv(base_nome, data_inicio, data_fim, cabecalho, linhas):
    """Devolve uma resposta Flask com um CSV pronto para download —
    separador ';' e BOM UTF-8 para abrir corretamente acentuado no Excel
    em português (que por padrão espera latin1/BOM, não UTF-8 puro)."""
    buffer = io.StringIO()
    buffer.write("﻿")
    escritor = csv.writer(buffer, delimiter=";")
    escritor.writerow(cabecalho)
    for linha in linhas:
        escritor.writerow(linha)

    nome_arquivo = _nome_arquivo(base_nome, data_inicio, data_fim, "csv")
    return Response(
        buffer.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


def exportar_xlsx(base_nome, data_inicio, data_fim, titulo, cabecalho, linhas):
    """Devolve uma resposta Flask com uma planilha .xlsx pronta para
    download, usando openpyxl (já é dependência do projeto — ver
    app/xlsx_preparo.py)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"])
    ws.append([])

    ws.append(cabecalho)
    for celula in ws[ws.max_row]:
        celula.font = Font(bold=True)

    for linha in linhas:
        ws.append(linha)

    for coluna in ws.columns:
        maior = max((len(str(c.value)) for c in coluna if c.value is not None), default=10)
        ws.column_dimensions[coluna[0].column_letter].width = min(maior + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = _nome_arquivo(base_nome, data_inicio, data_fim, "xlsx")
    return Response(
        buffer.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


def exportar_pdf(base_nome, data_inicio, data_fim, titulo, cabecalho, linhas, linhas_resumo=None):
    """Devolve uma resposta Flask com um PDF tabular pronto para download
    (reportlab, já usado em app/nfse_nacional.py e app/prontuario_pdf.py).
    `linhas_resumo`, se informado, é uma lista de strings mostrada acima da
    tabela (ex.: totais e médias) — pensado para impressão/anexo, não para
    reabrir e editar depois."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(
            f"Período: {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} — "
            f"gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            estilos["Normal"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    if linhas_resumo:
        for texto in linhas_resumo:
            elementos.append(Paragraph(texto, estilos["Normal"]))
        elementos.append(Spacer(1, 0.4 * cm))

    dados_tabela = [cabecalho] + [[("" if v is None else str(v)) for v in linha] for linha in linhas]
    tabela = Table(dados_tabela, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabela)

    if not linhas:
        elementos.append(Spacer(1, 0.4 * cm))
        elementos.append(Paragraph("Nenhum registro encontrado para este período/filtro.", estilos["Italic"]))

    doc.build(elementos)
    buffer.seek(0)

    nome_arquivo = _nome_arquivo(base_nome, data_inicio, data_fim, "pdf")
    return Response(
        buffer.read(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )
