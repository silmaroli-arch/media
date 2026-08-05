"""Extração de sugestões de preparo a partir de uma planilha Excel (.xlsx)
estruturada, como alternativa à extração heurística de PDF (ver
`app.pdf_preparo`) — quando a secretária/médico já organiza os dados do
preparo numa planilha com colunas fixas, a extração fica bem mais confiável
do que tentar interpretar texto solto de um PDF.

Formato esperado (uma linha por regra, cabeçalho na primeira linha):

    Tipo | Ação | Agrupador | Nome | Dias antes | Horas antes | Hora exata + dia antes

- Tipo: "Medicamento", "Alimento", "Aviso" ou "Exames / Procedimentos".
- Ação: depende do Tipo — "Suspender"/"Não suspender"/"Receituário" para
  Medicamento; "Suspender"/"Permitido (Sugestão de consumo)" para Alimento;
  "Proibido" para Exames/Procedimentos; qualquer texto para Aviso (o
  conteúdo do Aviso não depende da Ação, só do Tipo).
- Agrupador: opcional — categoria/classe (ex.: "medicamento
  antiplaquetário") ou, quando o "Nome" está em branco, o próprio nome do
  item (útil para linhas que descrevem uma classe inteira, sem marca
  específica, ex.: "antibióticos").
- Nome: nome do item (medicamento, alimento, exame/procedimento) ou o
  texto do aviso.
- Dias antes / Horas antes: prazo, num ou noutro, nunca os dois.
- Hora exata + dia antes: horário fixo do relógio, combinado com "Dias
  antes" quando o prazo é um horário certo num dia relativo ao exame (ex.:
  "pode comer até as 20:00 do dia anterior" -> Dias antes=1, Hora
  exata=20:00), diferente de um número de horas antes do exame.

Uma célula de "Tipo" em branco herda o Tipo (e a Ação) da linha anterior —
útil para uma sequência de avisos que não repete o rótulo em toda linha.

Cada aba da planilha é tratada como o preparo de um exame diferente —
`extrair_sugestoes_de_xlsx` retorna uma lista de sugestões, uma por aba,
no mesmo formato usado pela extração de PDF (compatível com o mesmo
formulário de revisão antes de salvar)."""
import re
import unicodedata

import openpyxl


def _normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto


def _texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def _inteiro(valor):
    """Converte para int quando possível — algumas planilhas guardam
    número como texto, ou como float (ex.: 12.0)."""
    if valor is None or valor == "":
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def _hora_hhmm(valor):
    """Normaliza a célula de hora para o formato 'HH:MM' — o Excel entrega
    um `datetime.time` quando a coluna está formatada como hora, mas
    aceita também texto solto (ex.: '16:00')."""
    if valor is None or valor == "":
        return None
    if hasattr(valor, "strftime"):
        return valor.strftime("%H:%M")
    texto = str(valor).strip()
    match = re.match(r"^(\d{1,2}):(\d{2})", texto)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return None


def _linhas_da_aba(planilha):
    """Gera (tipo, acao, agrupador, nome, dias_antes, horas_antes,
    hora_exata) para cada linha de dados, pulando o cabeçalho e linhas
    totalmente vazias, e herdando Tipo/Ação de uma linha anterior quando a
    célula de Tipo está em branco (continuação de uma mesma seção)."""
    ultimo_tipo, ultima_acao = None, None
    for i, linha in enumerate(planilha.iter_rows(values_only=True)):
        if i == 0:
            continue  # cabeçalho
        if not linha or all(c is None for c in linha):
            continue
        tipo_raw, acao_raw, agrupador, nome, dias_antes, horas_antes, hora_exata = (
            list(linha) + [None] * (7 - len(linha))
        )[:7]

        if _texto(tipo_raw):
            ultimo_tipo, ultima_acao = tipo_raw, acao_raw
        tipo_raw, acao_raw = ultimo_tipo, ultima_acao

        nome_texto = _texto(nome)
        agrupador_texto = _texto(agrupador)
        if not nome_texto and not agrupador_texto:
            continue

        yield (
            _normalizar(tipo_raw), _normalizar(acao_raw), agrupador_texto, nome_texto,
            _inteiro(dias_antes), _inteiro(horas_antes), _hora_hhmm(hora_exata),
        )


def _extrair_sugestao_da_aba(planilha, nome_aba):
    cortes = []
    medicamentos = []
    medicamentos_mantidos = []
    alimentos = []
    exames_anteriores = []
    informacoes_gerais = []

    for tipo, acao, agrupador, nome, dias_antes, horas_antes, hora_exata in _linhas_da_aba(planilha):
        if tipo == "medicamento":
            nome_item = nome or agrupador
            if not nome_item:
                continue
            if "nao suspender" in acao:
                medicamentos_mantidos.append({
                    "nome": nome_item,
                    "observacao": agrupador if (agrupador and nome) else None,
                })
            elif "receituario" in acao:
                # Instrução de dosagem com horário certo (ex.: sachê às
                # 16:00 do dia anterior) — vira um item de informação geral
                # com prazo calculável, em vez de um medicamento a suspender.
                informacoes_gerais.append({
                    "texto": nome_item,
                    "horas_antes": None,
                    "dias_antes": dias_antes,
                    "hora_exata": hora_exata,
                })
            else:
                if dias_antes is None:
                    continue
                medicamentos.append({
                    "nome": nome_item, "dias_antes": dias_antes,
                    "categoria": agrupador if (agrupador and nome) else None,
                })

        elif tipo == "alimento":
            nome_item = nome or agrupador
            if not nome_item:
                continue
            alimentos.append({
                "nome": nome_item,
                "permitido": "permitido" in acao,
                "horas_antes": horas_antes if dias_antes is None else None,
                "dias_antes": dias_antes,
            })

        elif "exame" in tipo or "procedimento" in tipo:
            if not nome:
                continue
            exames_anteriores.append({"nome": nome, "dias_antes": dias_antes})

        elif tipo == "aviso":
            texto_item = nome or agrupador
            if not texto_item:
                continue
            if "jejum" in texto_item.lower() and horas_antes:
                cortes.append({"descricao": texto_item, "horas_antes": horas_antes})
            elif horas_antes and not dias_antes:
                informacoes_gerais.append({
                    "texto": texto_item, "horas_antes": horas_antes,
                    "dias_antes": None, "hora_exata": None,
                })
            elif dias_antes and hora_exata:
                informacoes_gerais.append({
                    "texto": texto_item, "horas_antes": None,
                    "dias_antes": dias_antes, "hora_exata": hora_exata,
                })
            else:
                informacoes_gerais.append({
                    "texto": texto_item, "horas_antes": None,
                    "dias_antes": None, "hora_exata": None,
                })

    instrucoes = "\n".join(f"- {info['texto']}" for info in informacoes_gerais if info.get("texto"))

    return {
        "origem": "xlsx",
        "aba_nome": nome_aba,
        "nome_sugerido": None,
        "instrucoes": instrucoes,
        "cortes": cortes,
        "medicamentos": medicamentos,
        "medicamentos_mantidos": medicamentos_mantidos,
        "observacoes_medicamentos": None,
        "informacoes_gerais": informacoes_gerais,
        "alimentos": alimentos,
        "exames_anteriores": exames_anteriores,
    }


def extrair_sugestoes_de_xlsx(stream):
    """Lê a planilha e retorna uma lista de sugestões (uma por aba), no
    mesmo formato usado pela extração de PDF — cada uma pode ser revisada
    e salva como um modelo de preparo separado."""
    workbook = openpyxl.load_workbook(stream, data_only=True, read_only=True)
    return [
        _extrair_sugestao_da_aba(workbook[nome_aba], nome_aba)
        for nome_aba in workbook.sheetnames
    ]
