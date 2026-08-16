"""Teste de ponta a ponta de duas pontes entre o grupo de trabalho e
funcionalidades que já existiam do lado da equipe (Empresa/Clínica), BBP
seção 8:

- decisão nº 6 / telas 5.1.10-5.1.12: importação de modelo de preparo a
  partir de PDF/Excel, agora também dentro de um grupo (a extração em si
  não muda nada — só a tela de destino, que salva na clínica interna do
  grupo em vez de depender da clínica selecionada na sessão).
- decisão nº 5: quando um exame tem mais de um médico vinculado, qualquer
  um deles pode aprovar a resposta rascunhada pela IA antes dela ir para o
  paciente — a regra já existia em Exame.medico_pode_atender; só faltava
  uma tela de perguntas pendentes que buscasse pelo grupo (clínica interna),
  não pelas clínicas formais do usuário."""
import io
from datetime import datetime

from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import (
    Usuario, Grupo, GrupoMembro, GrupoPaciente, Paciente,
    Empresa, Clinica, ClinicaMembro, PreparoModelo, Exame, PerguntaPendente, FaqItem,
)

app = create_app()


def checar(nome, condicao):
    status = "OK" if condicao else "FALHOU"
    print(f"[{status}] {nome}")
    assert condicao, nome


def _construir_xlsx_teste():
    cabecalho = ["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"]
    wb = Workbook()
    aba1 = wb.active
    aba1.title = "Preparo A"
    for linha in [
        cabecalho,
        ["Medicamento", "Suspender", "medicamento antiplaquetário", "Ticlid", 10, None, None],
        ["Alimento", "Suspender", None, "frutas", 3, None, None],
        ["Exames / Procedimentos", "Proibido", None, "colonoscopia", 28, None, None],
        ["Aviso", "Intruções para IA", None, "JEJUM de 12 horas", None, 12, None],
    ]:
        aba1.append(linha)
    aba2 = wb.create_sheet("Preparo B")
    for linha in [cabecalho, ["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None]]:
        aba2.append(linha)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _construir_xlsx_uma_aba():
    wb = Workbook()
    wb.active.title = "Único preparo"
    wb.active.append(["Tipo", "Ação", "Agrupador", "Nome", "Dias antes", "Horas antes", "Hora exata + dia antes"])
    wb.active.append(["Aviso", "Intruções para IA", None, "Trazer o pedido médico", None, None, None])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


with app.app_context():
    for email in ("medico7.bbp@teste.com", "medico8.bbp@teste.com", "medico9.bbp@teste.com", "secretaria7.bbp@teste.com"):
        u = Usuario.query.filter_by(email=email).first()
        if u:
            GrupoMembro.query.filter_by(usuario_id=u.id).delete()
            db.session.delete(u)
    db.session.commit()

    medico_principal = Usuario(nome="Dr. Igor Passos", email="medico7.bbp@teste.com", tipo="medico",
                                cpf="345.987.612-14", crm_numero="5555", crm_uf="ES")
    medico_principal.set_senha("123456")
    medico_principal.definir_permissoes_padrao()
    medico_extra = Usuario(nome="Dra. Julia Rezende", email="medico8.bbp@teste.com", tipo="medico",
                            cpf="456.098.723-25", crm_numero="6666", crm_uf="ES")
    medico_extra.set_senha("123456")
    medico_extra.definir_permissoes_padrao()
    medico_fora = Usuario(nome="Dr. Kleber Matos", email="medico9.bbp@teste.com", tipo="medico",
                           cpf="567.109.834-36", crm_numero="7777", crm_uf="ES")
    medico_fora.set_senha("123456")
    medico_fora.definir_permissoes_padrao()
    secretaria = Usuario(nome="Laura Fontes", email="secretaria7.bbp@teste.com", tipo="secretaria",
                          cpf="678.210.945-47")
    secretaria.set_senha("123456")
    secretaria.definir_permissoes_padrao()
    db.session.add_all([medico_principal, medico_extra, medico_fora, secretaria])
    db.session.commit()
    medico_principal_id, medico_extra_id = medico_principal.id, medico_extra.id
    medico_fora_id, secretaria_id = medico_fora.id, secretaria.id

    empresa_teste = Empresa.query.filter_by(nome="Empresa Teste BBP").first()
    if not empresa_teste:
        empresa_teste = Empresa(nome="Empresa Teste BBP", status="ativa")
        db.session.add(empresa_teste)
        db.session.commit()
    clinica_teste = Clinica.query.filter_by(nome="Clínica Teste BBP", empresa_id=empresa_teste.id).first()
    if not clinica_teste:
        clinica_teste = Clinica(nome="Clínica Teste BBP", empresa_id=empresa_teste.id)
        db.session.add(clinica_teste)
        db.session.commit()
    for uid in (medico_principal_id, medico_extra_id, medico_fora_id, secretaria_id):
        if not ClinicaMembro.query.filter_by(clinica_id=clinica_teste.id, usuario_id=uid).first():
            db.session.add(ClinicaMembro(clinica_id=clinica_teste.id, usuario_id=uid))
    db.session.commit()

    paciente = Paciente(
        empresa_id=None, nome="Paciente Import BBP", cpf="789.321.056-58",
        telefone="(27) 97777-0000", cep="29000-000", rua="Rua Import", numero="1",
        bairro="Centro", cidade="Vitória", uf="ES",
    )
    db.session.add(paciente)
    db.session.commit()
    paciente_id = paciente.id


client_medico = app.test_client()
client_medico_extra = app.test_client()
client_medico_fora = app.test_client()
client_secretaria = app.test_client()


def login(client, cpf, senha):
    return client.post("/login", data={"identificador": cpf, "senha": senha}, follow_redirects=True)


login(client_medico, "345.987.612-14", "123456")
login(client_medico_extra, "456.098.723-25", "123456")
login(client_medico_fora, "567.109.834-36", "123456")
login(client_secretaria, "678.210.945-47", "123456")

client_medico.post("/grupos/novo", data={"nome": "Grupo Import BBP"}, follow_redirects=True)
with app.app_context():
    grupo = Grupo.query.filter_by(nome="Grupo Import BBP").order_by(Grupo.id.desc()).first()
    grupo_id = grupo.id

for cpf in ("456.098.723-25", "678.210.945-47"):
    r = client_medico.post(f"/grupos/{grupo_id}/convidar", data={"cpf": cpf}, follow_redirects=True)

with app.app_context():
    from app.models import GrupoConvite
    convite_extra_id = GrupoConvite.query.filter_by(grupo_id=grupo_id, usuario_convidado_id=medico_extra_id).first().id
    convite_secretaria_id = GrupoConvite.query.filter_by(grupo_id=grupo_id, usuario_convidado_id=secretaria_id).first().id

client_medico_extra.post(f"/grupos/convites/{convite_extra_id}/responder", data={"decisao": "aprovar"}, follow_redirects=True)
client_secretaria.post(f"/grupos/convites/{convite_secretaria_id}/responder", data={"decisao": "aprovar"}, follow_redirects=True)

client_medico.post(f"/grupos/{grupo_id}/pacientes/novo", data={
    "etapa": "salvar", "paciente_id": str(paciente_id), "grupos_ids": str(grupo_id),
}, follow_redirects=True)

# =============== Importação de PDF/Excel dentro do grupo ===============

r = client_medico_fora.get(f"/grupos/{grupo_id}/preparo-modelos/importar-xlsx", follow_redirects=True)
checar("Quem não participa do grupo não acessa a importação", "Você não participa deste grupo" in r.get_data(as_text=True))

r = client_secretaria.get(f"/grupos/{grupo_id}/preparo-modelos/importar-xlsx", follow_redirects=True)
checar("Secretária (não é médico) não pode importar modelo de preparo", "Somente usuários do tipo Médico" in r.get_data(as_text=True))

r = client_medico.get(f"/grupos/{grupo_id}/preparo-modelos/pdf-para-excel", follow_redirects=True)
checar("Tela de gerar Excel a partir de PDF carrega para o médico do grupo", r.status_code == 200)

# Planilha com 2 abas -> tela de escolha -> revisão pré-preenchida -> salvar.
r = client_medico.post(
    f"/grupos/{grupo_id}/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (_construir_xlsx_teste(), "teste_preparo.xlsx")},
    content_type="multipart/form-data", follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com 2 abas leva à tela de escolher qual aba importar primeiro (dentro do grupo)",
       "Preparo A" in texto and "Preparo B" in texto)

r = client_medico.post(f"/grupos/{grupo_id}/preparo-modelos/importar-xlsx/escolher", data={"indice": "0"}, follow_redirects=True)
texto = r.get_data(as_text=True)
checar("Ao escolher a aba, o formulário de novo modelo do GRUPO vem pré-preenchido com os dados extraídos",
       "Ticlid" in texto and "colonoscopia" in texto)

r = client_medico.post(f"/grupos/{grupo_id}/preparo-modelos/novo", data={
    "nome": "Preparo Importado BBP",
    "instrucoes": "Instruções de teste.",
    "corte_descricao[]": ["JEJUM de 12 horas"],
    "corte_horas[]": ["12"],
    "medicamento_nome[]": ["Ticlid"],
    "medicamento_categoria[]": ["medicamento antiplaquetário"],
    "medicamento_dias[]": ["10"],
    "medicamento_obs[]": [""],
    "alimento_nome[]": ["frutas"],
    "alimento_tipo[]": ["proibido"],
    "alimento_horas[]": [""],
    "alimento_dias[]": ["3"],
    "exame_anterior_nome[]": ["colonoscopia"],
    "exame_anterior_dias[]": ["28"],
}, follow_redirects=True)
checar("Modelo importado é salvo com sucesso no grupo", "cadastrado com sucesso" in r.get_data(as_text=True))

with app.app_context():
    grupo_local = Grupo.query.filter_by(nome="Grupo Import BBP").first()
    grupo_local_id = grupo_local.id
    modelo_importado = PreparoModelo.query.filter_by(grupo_id=grupo_local_id, nome="Preparo Importado BBP").first()
    checar("Modelo importado existe", modelo_importado is not None)
    checar("Modelo importado tem o corte de jejum calculável (12h antes)", any(c.horas_antes == 12 for c in modelo_importado.cortes))
    checar("Modelo importado tem o medicamento com categoria",
           any(m.medicamento.nome == "Ticlid" and m.medicamento.categoria == "medicamento antiplaquetário" for m in modelo_importado.medicamentos_suspensos))
    checar("Modelo importado tem o alimento com prazo em dias", any(a.nome == "frutas" and a.dias_antes == 3 for a in modelo_importado.alimentos))
    checar("Modelo importado tem o exame anterior proibido", any(e.nome == "colonoscopia" and e.dias_antes == 28 for e in modelo_importado.exames_anteriores_proibidos))
    modelo_importado_id = modelo_importado.id

# Planilha com uma aba só vai direto para a revisão, sem tela de escolha.
r = client_medico.post(
    f"/grupos/{grupo_id}/preparo-modelos/importar-xlsx",
    data={"arquivo_xlsx": (_construir_xlsx_uma_aba(), "teste_uma_aba.xlsx")},
    content_type="multipart/form-data", follow_redirects=True,
)
texto = r.get_data(as_text=True)
checar("Planilha com uma aba só vai direto para a tela de revisão do grupo (sem escolher aba)",
       "Trazer o pedido médico" in texto and "Novo modelo de preparo" in texto)


# =============== Aprovação de dúvida com múltiplos médicos vinculados ===============

r = client_medico.post(f"/grupos/{grupo_id}/exames/novo", data={
    "nome": "Exame Multi-Medico BBP", "preparo_modelo_id": str(modelo_importado_id),
}, follow_redirects=True)
checar("Exame com o modelo importado cadastrado", "cadastrado com sucesso" in r.get_data(as_text=True))

with app.app_context():
    grupo_local = Grupo.query.filter_by(nome="Grupo Import BBP").first()
    exame = Exame.query.filter_by(grupo_id=grupo_local.id, nome="Exame Multi-Medico BBP").first()
    checar("Exame criado com o médico principal", exame.medico_id == medico_principal_id)
    # Vincula um segundo médico ao exame (BBP: exame pode ter múltiplos médicos).
    medico_extra_obj = Usuario.query.get(medico_extra_id)
    exame.medicos_extra.append(medico_extra_obj)
    db.session.commit()
    exame_id = exame.id

    pergunta_ia = PerguntaPendente(
        grupo_id=grupo_local.id, paciente_id=paciente_id, exame_id=exame_id,
        pergunta="Posso tomar água antes do exame?", status="aguardando_aprovacao",
        resposta_sugerida_ia="Sim, água em pequena quantidade é permitida até 2 horas antes.",
    )
    db.session.add(pergunta_ia)
    db.session.commit()
    pergunta_ia_id = pergunta_ia.id

# O médico fora do grupo nem vê a tela.
r = client_medico_fora.get(f"/grupos/{grupo_id}/perguntas", follow_redirects=True)
checar("Médico de fora do grupo não acessa a tela de perguntas", "Você não participa deste grupo" in r.get_data(as_text=True))

# Ambos os médicos vinculados ao exame (principal e extra) veem a pergunta.
r = client_medico.get(f"/grupos/{grupo_id}/perguntas", follow_redirects=True)
checar("Médico principal vê a pergunta aguardando aprovação", "Posso tomar água antes do exame?" in r.get_data(as_text=True))
r = client_medico_extra.get(f"/grupos/{grupo_id}/perguntas", follow_redirects=True)
checar("Médico EXTRA (não principal) também vê a pergunta aguardando aprovação (BBP decisão nº 5)",
       "Posso tomar água antes do exame?" in r.get_data(as_text=True))

# O médico EXTRA (não principal) consegue aprovar — é essa a regra da decisão nº 5.
r = client_medico_extra.post(f"/grupos/{grupo_id}/perguntas/{pergunta_ia_id}/responder", data={
    "resposta": "Sim, água em pequena quantidade é permitida até 2 horas antes do exame.",
}, follow_redirects=True)
checar("Médico EXTRA consegue aprovar a resposta da IA (qualquer médico vinculado ao exame pode aprovar)",
       "Resposta salva" in r.get_data(as_text=True))

with app.app_context():
    pergunta_atualizada = PerguntaPendente.query.filter_by(id=pergunta_ia_id).first()
    checar("Pergunta ficou com status 'respondida'", pergunta_atualizada.status == "respondida")
    checar("Pergunta registra quem respondeu (o médico extra)", pergunta_atualizada.respondida_por == "Dra. Julia Rezende")
    faq = FaqItem.query.filter_by(exame_id=exame_id, pergunta="Posso tomar água antes do exame?").first()
    checar("Resposta aprovada entrou na base de conhecimento (FaqItem) com o grupo_id do grupo",
           faq is not None and faq.grupo_id == grupo_local_id)

    # Segunda pergunta, desta vez sem exame associado (geral) e sem rascunho da IA.
    pergunta_geral = PerguntaPendente(
        grupo_id=grupo_local_id, paciente_id=paciente_id, exame_id=None,
        pergunta="Qual o telefone da clínica?", status="pendente",
    )
    db.session.add(pergunta_geral)
    db.session.commit()
    pergunta_geral_id = pergunta_geral.id

with app.app_context():
    # Um segundo exame, do médico extra desta vez, para provar que o médico
    # principal NÃO pode responder perguntas de um exame que não é dele nem
    # tem como "extra".
    exame2 = Exame(
        grupo_id=grupo_local_id, criado_por_id=medico_extra_id,
        medico_id=medico_extra_id, nome="Exame Só Médico Extra BBP",
        associado=True, medico_confirmado=True,
    )
    db.session.add(exame2)
    db.session.commit()
    exame2_id = exame2.id
    pergunta_exame_alheio = PerguntaPendente(
        grupo_id=grupo_local_id, paciente_id=paciente_id, exame_id=exame2_id,
        pergunta="Preciso jejuar?", status="pendente",
    )
    db.session.add(pergunta_exame_alheio)
    db.session.commit()
    pergunta_exame_alheio_id = pergunta_exame_alheio.id

r = client_medico.post(f"/grupos/{grupo_id}/perguntas/{pergunta_exame_alheio_id}/responder", data={
    "resposta": "Não precisa.",
}, follow_redirects=True)
checar("Médico principal NÃO pode responder pergunta de exame que não é dele nem tem como extra",
       "Você só pode responder perguntas sobre os seus próprios exames" in r.get_data(as_text=True))
with app.app_context():
    checar("Pergunta de exame alheio continua sem resposta",
           PerguntaPendente.query.get(pergunta_exame_alheio_id).status == "pendente")

# A secretaria do grupo (com perm_pacientes) responde a pergunta geral (sem exame).
r = client_secretaria.post(f"/grupos/{grupo_id}/perguntas/{pergunta_geral_id}/responder", data={
    "resposta": "(27) 3333-0000.",
}, follow_redirects=True)
checar("Secretária do grupo consegue responder pergunta geral (sem exame associado)",
       "Resposta salva" in r.get_data(as_text=True))

print("\nTodas as verificações de importação PDF/Excel e aprovação multi-médico dentro do grupo passaram.")
